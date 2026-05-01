"""高级模块执行器 - advanced_excel"""
from .base import ModuleExecutor, ExecutionContext, ModuleResult, register_executor
from .type_utils import to_int, to_float, parse_search_region
import asyncio
import os
import re


@register_executor
class ReadExcelExecutor(ModuleExecutor):
    """Excel文件读取模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "read_excel"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import openpyxl
        from app.api.data_assets import get_asset_by_name
        
        file_name = context.resolve_value(config.get('fileName', ''))
        sheet_name = context.resolve_value(config.get('sheetName', ''))
        read_mode = context.resolve_value(config.get('readMode', 'cell'))  # 支持变量引用
        cell_address = context.resolve_value(config.get('cellAddress', ''))
        row_index = to_int(config.get('rowIndex', 1), 1, context)
        column_index = context.resolve_value(config.get('columnIndex', ''))
        start_cell = context.resolve_value(config.get('startCell', ''))
        end_cell = context.resolve_value(config.get('endCell', ''))
        variable_name = config.get('variableName', '')
        start_row = to_int(config.get('startRow', 2), 2, context)
        start_col = context.resolve_value(config.get('startCol', ''))
        
        if not file_name:
            return ModuleResult(success=False, error="请选择要读取的Excel文件")
        
        if not variable_name:
            return ModuleResult(success=False, error="请指定存储变量名")
        
        asset = get_asset_by_name(file_name)
        if not asset:
            return ModuleResult(success=False, error=f"文件 '{file_name}' 不存在")
        
        file_path = asset['path']
        is_xls = file_path.lower().endswith('.xls')
        
        try:
            # 使用线程池执行同步Excel操作
            loop = asyncio.get_running_loop()
            if is_xls:
                result, result_type = await loop.run_in_executor(
                    None, self._read_xls, file_path, sheet_name, read_mode, 
                    cell_address, row_index, column_index, start_cell, end_cell, start_row, start_col
                )
            else:
                result, result_type = await loop.run_in_executor(
                    None, self._read_xlsx, file_path, sheet_name, read_mode,
                    cell_address, row_index, column_index, start_cell, end_cell, start_row, start_col
                )
            
            context.set_variable(variable_name, result)
            
            display_content = str(result)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(
                success=True,
                message=f"已读取Excel数据: {display_content}",
                data={'value': result, 'type': result_type, 'file': file_name}
            )
        except Exception as e:
            return ModuleResult(success=False, error=f"读取Excel失败: {str(e)}")
    
    def _read_xlsx(self, file_path, sheet_name, read_mode, cell_address, row_index, 
                   column_index, start_cell, end_cell, start_row, start_col):
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise Exception(f"工作表 '{sheet_name}' 不存在")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        result = None
        result_type = 'unknown'
        
        if read_mode == 'cell':
            if not cell_address:
                wb.close()
                raise Exception("单元格模式需要指定单元格地址")
            cell = ws[cell_address]
            result = cell.value if cell.value is not None else ''
            result_type = 'cell'
        
        elif read_mode == 'row':
            if row_index is None or row_index < 1:
                wb.close()
                raise Exception("行模式需要指定有效的行号")
            row_data = []
            start_col_idx = 1
            if start_col:
                if isinstance(start_col, str) and start_col.isalpha():
                    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
                else:
                    start_col_idx = int(start_col)
            for idx, cell in enumerate(ws[row_index], start=1):
                if idx >= start_col_idx:
                    row_data.append(cell.value if cell.value is not None else '')
            result = row_data
            result_type = 'array'
        
        elif read_mode == 'column':
            if not column_index:
                wb.close()
                raise Exception("列模式需要指定列号或列字母")
            col_data = []
            col_idx = column_index
            if isinstance(col_idx, str) and col_idx.isalpha():
                col_idx = openpyxl.utils.column_index_from_string(col_idx)
            else:
                col_idx = int(col_idx)
            for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
                col_data.append(row[0].value if row[0].value is not None else '')
            result = col_data
            result_type = 'array'
        
        elif read_mode == 'range':
            if not start_cell or not end_cell:
                wb.close()
                raise Exception("范围模式需要指定起始和结束单元格")
            range_data = []
            for row in ws[f"{start_cell}:{end_cell}"]:
                row_data = [cell.value if cell.value is not None else '' for cell in row]
                range_data.append(row_data)
            result = range_data
            result_type = 'matrix'
        
        wb.close()
        return result, result_type
    
    def _read_xls(self, file_path, sheet_name, read_mode, cell_address, row_index,
                  column_index, start_cell, end_cell, start_row, start_col):
        import xlrd
        
        wb = xlrd.open_workbook(file_path)
        
        if sheet_name:
            if sheet_name not in wb.sheet_names():
                raise Exception(f"工作表 '{sheet_name}' 不存在")
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
        
        result = None
        result_type = 'unknown'
        
        if read_mode == 'cell':
            if not cell_address:
                raise Exception("单元格模式需要指定单元格地址")
            col_idx, row_idx = self._parse_cell_address(cell_address)
            value = ws.cell_value(row_idx, col_idx)
            result = value if value != '' else ''  # xlrd空单元格返回空字符串
            result_type = 'cell'
        
        elif read_mode == 'row':
            if row_index is None or row_index < 1:
                raise Exception("行模式需要指定有效的行号")
            row_data = ws.row_values(row_index - 1)
            # 将空值转换为空字符串
            row_data = [v if v != '' else '' for v in row_data]
            if start_col:
                start_col_idx = 0
                if isinstance(start_col, str) and start_col.isalpha():
                    start_col_idx = self._col_letter_to_index(start_col)
                else:
                    start_col_idx = int(start_col) - 1
                row_data = row_data[start_col_idx:]
            result = row_data
            result_type = 'array'
        
        elif read_mode == 'column':
            if not column_index:
                raise Exception("列模式需要指定列号或列字母")
            col_idx = column_index
            if isinstance(col_idx, str) and col_idx.isalpha():
                col_idx = self._col_letter_to_index(col_idx)
            else:
                col_idx = int(col_idx) - 1
            col_data = ws.col_values(col_idx, start_rowx=start_row - 1)
            # 将空值转换为空字符串
            col_data = [v if v != '' else '' for v in col_data]
            result = col_data
            result_type = 'array'
        
        elif read_mode == 'range':
            if not start_cell or not end_cell:
                raise Exception("范围模式需要指定起始和结束单元格")
            start_col_idx, start_row_idx = self._parse_cell_address(start_cell)
            end_col_idx, end_row_idx = self._parse_cell_address(end_cell)
            range_data = []
            for r in range(start_row_idx, end_row_idx + 1):
                row_data = []
                for c in range(start_col_idx, end_col_idx + 1):
                    value = ws.cell_value(r, c)
                    row_data.append(value if value != '' else '')
                range_data.append(row_data)
            result = range_data
            result_type = 'matrix'
        
        return result, result_type
    
    def _col_letter_to_index(self, col_str: str) -> int:
        result = 0
        for c in col_str.upper():
            result = result * 26 + (ord(c) - ord('A') + 1)
        return result - 1
    
    def _parse_cell_address(self, address: str) -> tuple:
        col_str, row_str = '', ''
        for c in address:
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        return self._col_letter_to_index(col_str), int(row_str) - 1