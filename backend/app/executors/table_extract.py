"""表格数据提取模块"""
import os
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .base import (
    ModuleExecutor,
    ExecutionContext,
    ModuleResult,
    register_executor,
)


@register_executor
class ExtractTableDataExecutor(ModuleExecutor):
    """表格数据提取模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "extract_table_data"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        """
        提取网页表格数据
        
        配置参数:
        - tableSelector: 表格选择器
        - variableName: 存储变量名（存储二维列表）
        - exportToExcel: 是否导出为Excel
        - excelPath: Excel文件保存路径（可选）
        - includeHeader: 是否包含表头
        - headerRow: 表头行索引（默认0，第一行）
        """
        
        table_selector = context.resolve_value(config.get('tableSelector', ''))
        variable_name = config.get('variableName', 'table_data')
        export_to_excel = config.get('exportToExcel', False)
        excel_path = context.resolve_value(config.get('excelPath', ''))
        include_header = config.get('includeHeader', True)
        header_row = int(config.get('headerRow', 0))
        
        if not table_selector:
            return ModuleResult(success=False, error="表格选择器不能为空")
        
        page = await context.get_current_frame()
        if not page:
            return ModuleResult(success=False, error="未找到活动页面")
        
        try:
            # 等待表格元素出现
            try:
                await page.wait_for_selector(table_selector, timeout=10000)
            except Exception as e:
                return ModuleResult(
                    success=False,
                    error=f"未找到表格元素: {table_selector}"
                )
            
            # 使用 Playwright 的 locator API 提取表格数据（更可靠）
            print(f"[ExtractTable] 使用 locator API 提取表格，选择器: {table_selector}")
            
            try:
                # 获取表格元素
                table_locator = page.locator(table_selector).first
                
                # 检查元素是否存在
                is_visible = await table_locator.is_visible()
                if not is_visible:
                    return ModuleResult(success=False, error=f"表格元素不可见: {table_selector}")
                
                # 获取元素的标签名
                tag_name = await table_locator.evaluate("el => el.tagName")
                print(f"[ExtractTable] 元素标签名: {tag_name}")
                
                # 如果不是 TABLE 元素，尝试查找最近的 table
                if tag_name != 'TABLE':
                    print(f"[ExtractTable] 元素不是TABLE，尝试查找最近的table")
                    table_locator = table_locator.locator('xpath=ancestor-or-self::table').first
                    is_table_visible = await table_locator.is_visible()
                    if not is_table_visible:
                        return ModuleResult(success=False, error="选择的元素不在表格内")
                
                # 获取所有行
                rows_locator = table_locator.locator('tr')
                row_count = await rows_locator.count()
                print(f"[ExtractTable] 找到 {row_count} 行")
                
                if row_count == 0:
                    return ModuleResult(success=False, error="表格中没有行")
                
                # 提取每一行的数据
                table_data = []
                max_columns = 0
                
                for i in range(row_count):
                    row_locator = rows_locator.nth(i)
                    
                    # 获取 th 和 td 单元格
                    cells_locator = row_locator.locator('th, td')
                    cell_count = await cells_locator.count()
                    
                    if cell_count == 0:
                        continue
                    
                    row_data = []
                    for j in range(cell_count):
                        cell_locator = cells_locator.nth(j)
                        cell_text = await cell_locator.inner_text()
                        row_data.append(cell_text.strip())
                    
                    if row_data:
                        table_data.append(row_data)
                        max_columns = max(max_columns, len(row_data))
                
                print(f"[ExtractTable] 成功提取 {len(table_data)} 行数据，最大列数: {max_columns}")
                
                if not table_data:
                    return ModuleResult(success=False, error="表格为空或未找到数据")
                
                # 设置结果变量
                row_count = len(table_data)
                column_count = max_columns
                
            except Exception as e:
                print(f"[ExtractTable] locator API 失败: {str(e)}")
                import traceback
                traceback.print_exc()
                return ModuleResult(success=False, error=f"提取表格失败: {str(e)}")
            
            # 保存到变量
            if variable_name:
                context.set_variable(variable_name, table_data)
            
            # 导出为Excel
            if export_to_excel:
                if not excel_path:
                    # 如果没有指定路径，使用默认路径
                    excel_path = os.path.join(os.getcwd(), 'table_data.xlsx')
                
                try:
                    # 确保目录存在
                    excel_dir = os.path.dirname(excel_path)
                    if excel_dir and not os.path.exists(excel_dir):
                        os.makedirs(excel_dir, exist_ok=True)
                    
                    # 创建工作簿
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "表格数据"
                    
                    # 写入数据
                    for row_idx, row_data in enumerate(table_data, start=1):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                            
                            # 如果是表头行，设置样式
                            if include_header and row_idx == (header_row + 1):
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # 自动调整列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # 保存文件
                    wb.save(excel_path)
                    
                    return ModuleResult(
                        success=True,
                        message=f"成功提取表格数据（{row_count}行 x {column_count}列），已导出到: {excel_path}",
                        data={
                            'rowCount': row_count,
                            'columnCount': column_count,
                            'excelPath': excel_path,
                            'tableData': table_data
                        }
                    )
                    
                except Exception as e:
                    return ModuleResult(
                        success=False,
                        error=f"导出Excel失败: {str(e)}"
                    )
            
            return ModuleResult(
                success=True,
                message=f"成功提取表格数据（{row_count}行 x {column_count}列）",
                data={
                    'rowCount': row_count,
                    'columnCount': column_count,
                    'tableData': table_data
                }
            )
            
        except Exception as e:
            return ModuleResult(success=False, error=f"提取表格数据失败: {str(e)}")
