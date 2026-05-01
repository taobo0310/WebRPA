"""高级模块执行器实现 - 异步版本"""
import asyncio
import base64
import json
import random
import time
from pathlib import Path

from .base import (
    ModuleExecutor,
    ExecutionContext,
    ModuleResult,
    register_executor,
    pw_wait_for_element,

    format_selector
)
from .type_utils import to_int, to_float, parse_search_region
from ..utils.jsonpath_parser import parse_jsonpath


@register_executor
class ApiRequestExecutor(ModuleExecutor):
    """API请求模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "api_request"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import httpx
        
        request_url = context.resolve_value(config.get('requestUrl', ''))
        request_method = context.resolve_value(config.get('requestMethod', 'GET')).upper()  # 支持变量引用
        request_headers_str = context.resolve_value(config.get('requestHeaders', ''))
        request_body_str = context.resolve_value(config.get('requestBody', ''))
        variable_name = config.get('variableName', '')
        request_timeout = to_int(config.get('requestTimeout', 30), 30, context)
        
        if not request_url:
            return ModuleResult(success=False, error="请求地址不能为空")
        
        try:
            headers = {}
            if request_headers_str:
                try:
                    headers = json.loads(request_headers_str)
                except json.JSONDecodeError as e:
                    return ModuleResult(success=False, error=f"请求头JSON格式错误: {str(e)}")
            
            body = None
            if request_body_str:
                try:
                    body = json.loads(request_body_str)
                except json.JSONDecodeError:
                    body = request_body_str
            
            async with httpx.AsyncClient(timeout=request_timeout) as client:
                response = await client.request(
                    method=request_method,
                    url=request_url,
                    headers=headers,
                    json=body if isinstance(body, dict) else None,
                    data=body if isinstance(body, str) else None,
                )
            
            try:
                response_data = response.json()
            except:
                response_data = response.text
            
            if variable_name:
                context.set_variable(variable_name, response_data)
            
            display_content = str(response_data)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(
                success=True,
                message=f"请求成功 ({response.status_code}): {display_content}",
                data={'status_code': response.status_code, 'response': response_data}
            )
        
        except httpx.TimeoutException:
            return ModuleResult(success=False, error=f"请求超时 ({request_timeout}秒)")
        except httpx.ConnectError:
            return ModuleResult(success=False, error="无法连接到服务器，请检查网络和URL")
        except Exception as e:
            return ModuleResult(success=False, error=f"API请求失败: {str(e)}")


@register_executor
class JsonParseExecutor(ModuleExecutor):
    """JSON解析模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "json_parse"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        source_variable = config.get('sourceVariable', '')
        json_path = context.resolve_value(config.get('jsonPath', ''))  # 支持变量引用
        variable_name = config.get('variableName', '')
        column_name = config.get('columnName', '')
        
        if not source_variable:
            return ModuleResult(success=False, error="源数据变量不能为空")
        
        if not json_path:
            return ModuleResult(success=False, error="JSONPath表达式不能为空")
        
        source_data = context.get_variable(source_variable)
        if source_data is None:
            return ModuleResult(success=False, error=f"变量 '{source_variable}' 不存在")
        
        if isinstance(source_data, str):
            try:
                source_data = json.loads(source_data)
            except json.JSONDecodeError as e:
                return ModuleResult(success=False, error=f"源数据不是有效的JSON: {str(e)}")
        
        try:
            result = parse_jsonpath(source_data, json_path)
            
            if variable_name:
                context.set_variable(variable_name, result)
            
            if column_name:
                context.add_data_value(column_name, result)
            
            display_content = str(result)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(success=True, message=f"解析成功: {display_content}", data=result)
        
        except Exception as e:
            return ModuleResult(success=False, error=f"JSON解析失败: {str(e)}")


@register_executor
class Base64Executor(ModuleExecutor):
    """Base64处理模块执行器"""

    @property
    def module_type(self) -> str:
        return "base64"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        operation = context.resolve_value(config.get("operation", "encode"))  # 支持变量引用
        variable_name = config.get("variableName", "")

        try:
            if operation == "encode":
                input_text = context.resolve_value(config.get("inputText", ""))
                if not input_text:
                    return ModuleResult(success=False, error="输入文本不能为空")

                result = base64.b64encode(input_text.encode("utf-8")).decode("utf-8")
                if variable_name:
                    context.set_variable(variable_name, result)

                display = result[:50] + "..." if len(result) > 50 else result
                return ModuleResult(success=True, message=f"编码成功: {display}", data=result)

            elif operation == "decode":
                input_base64 = context.resolve_value(config.get("inputBase64", ""))
                if not input_base64:
                    return ModuleResult(success=False, error="Base64字符串不能为空")

                if "," in input_base64:
                    input_base64 = input_base64.split(",", 1)[1]

                result = base64.b64decode(input_base64).decode("utf-8")
                if variable_name:
                    context.set_variable(variable_name, result)

                display = result[:50] + "..." if len(result) > 50 else result
                return ModuleResult(success=True, message=f"解码成功: {display}", data=result)

            elif operation == "file_to_base64":
                file_path = context.resolve_value(config.get("filePath", ""))
                if not file_path:
                    return ModuleResult(success=False, error="文件路径不能为空")

                path = Path(file_path)
                if not path.exists():
                    return ModuleResult(success=False, error=f"文件不存在: {file_path}")

                with open(path, "rb") as f:
                    file_data = f.read()

                result = base64.b64encode(file_data).decode("utf-8")

                ext = path.suffix.lower()
                mime_types = {
                    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                    ".gif": "image/gif", ".webp": "image/webp", ".svg": "image/svg+xml",
                    ".pdf": "application/pdf", ".txt": "text/plain",
                    ".json": "application/json", ".xml": "application/xml",
                }
                mime_type = mime_types.get(ext, "application/octet-stream")
                data_url = f"data:{mime_type};base64,{result}"

                if variable_name:
                    context.set_variable(variable_name, data_url)

                return ModuleResult(
                    success=True,
                    message=f"文件转换成功: {path.name} ({len(file_data)} 字节)",
                    data=data_url,
                )

            elif operation == "base64_to_file":
                input_base64 = context.resolve_value(config.get("inputBase64", ""))
                output_path = context.resolve_value(config.get("outputPath", ""))
                file_name = context.resolve_value(config.get("fileName", "output.bin"))

                if not input_base64:
                    return ModuleResult(success=False, error="Base64字符串不能为空")
                if not output_path:
                    return ModuleResult(success=False, error="保存路径不能为空")

                if "," in input_base64:
                    input_base64 = input_base64.split(",", 1)[1]

                file_data = base64.b64decode(input_base64)

                output_dir = Path(output_path)
                output_dir.mkdir(parents=True, exist_ok=True)

                full_path = output_dir / file_name
                with open(full_path, "wb") as f:
                    f.write(file_data)

                result_path = str(full_path)
                if variable_name:
                    context.set_variable(variable_name, result_path)

                return ModuleResult(
                    success=True,
                    message=f"文件保存成功: {full_path} ({len(file_data)} 字节)",
                    data=result_path,
                )

            else:
                return ModuleResult(success=False, error=f"未知操作类型: {operation}")

        except Exception as e:
            return ModuleResult(success=False, error=f"Base64处理失败: {str(e)}")


@register_executor
class SelectDropdownExecutor(ModuleExecutor):
    """下拉框选择模块执行器"""

    @property
    def module_type(self) -> str:
        return "select_dropdown"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get('selector', ''))
        select_by = context.resolve_value(config.get('selectBy', 'value'))  # 支持变量引用
        value = context.resolve_value(config.get('value', ''))
        # 获取超时配置（秒），转换为毫秒传给 Playwright
        timeout_ms = to_int(config.get('timeout', 30), 30, context) * 1000
        
        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            
            # 先等待页面加载完成
            # 处理超时参数：0 表示不限制超时，None 表示使用 Playwright 默认超时
            wait_timeout = None if timeout_ms == 0 else timeout_ms
            
            try:
                await context.page.wait_for_load_state('domcontentloaded', timeout=wait_timeout)
            except:
                pass
            
            await pw_wait_for_element(context.page, selector, state='visible', timeout=wait_timeout)
            element = context.page.locator(format_selector(selector))

            if select_by == 'value':
                await element.select_option(value=value)
            elif select_by == 'label':
                await element.select_option(label=value)
            elif select_by == 'index':
                await element.select_option(index=int(value))
            
            return ModuleResult(success=True, message=f"已选择: {value}")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"选择下拉框失败: {str(e)}")


@register_executor
class SetCheckboxExecutor(ModuleExecutor):
    """设置复选框模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "set_checkbox"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get('selector', ''))
        checked_raw = config.get('checked', True)
        # 支持变量引用
        if isinstance(checked_raw, str):
            checked_raw = context.resolve_value(checked_raw)
        checked = checked_raw in [True, 'true', 'True', '1', 1]
        
        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            await pw_wait_for_element(context.page, selector, state='visible')
            element = context.page.locator(format_selector(selector))
            
            if checked:
                await element.check()
            else:
                await element.uncheck()
            
            return ModuleResult(success=True, message=f"复选框已{'勾选' if checked else '取消勾选'}")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"设置复选框失败: {str(e)}")


@register_executor
class DragElementExecutor(ModuleExecutor):
    """拖拽元素模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "drag_element"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        source_selector = context.resolve_value(config.get('sourceSelector', ''))
        target_selector = context.resolve_value(config.get('targetSelector', ''))
        target_position = config.get('targetPosition')
        
        if not source_selector:
            return ModuleResult(success=False, error="源元素选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            source = context.page.locator(source_selector)
            
            if target_selector:
                target = context.page.locator(target_selector)
                await source.drag_to(target)
            elif target_position:
                box = await source.bounding_box()
                if box:
                    start_x = box['x'] + box['width'] / 2
                    start_y = box['y'] + box['height'] / 2
                    end_x = target_position.get('x', start_x)
                    end_y = target_position.get('y', start_y)
                    
                    await context.page.mouse.move(start_x, start_y)
                    await context.page.mouse.down()
                    await context.page.mouse.move(end_x, end_y, steps=10)
                    await context.page.mouse.up()
            
            return ModuleResult(success=True, message="拖拽完成")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"拖拽元素失败: {str(e)}")


@register_executor
class ScrollPageExecutor(ModuleExecutor):
    """滚动页面模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "scroll_page"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        direction = context.resolve_value(config.get('direction', 'down'))  # 支持变量引用
        distance = to_int(config.get('distance', 500), 500, context)
        selector = context.resolve_value(config.get('selector', ''))
        scroll_mode = context.resolve_value(config.get('scrollMode', 'auto'))  # 支持变量引用
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            
            delta_x = 0
            delta_y = 0
            
            if direction == 'down':
                delta_y = distance
            elif direction == 'up':
                delta_y = -distance
            elif direction == 'right':
                delta_x = distance
            elif direction == 'left':
                delta_x = -distance
            
            # 使用鼠标滚轮模拟滚动（对抖音等虚拟滚动页面更有效）
            if scroll_mode == 'wheel' or scroll_mode == 'auto':
                try:
                    if selector:
                        # 先定位到元素中心
                        element = context.page.locator(format_selector(selector)).first
                        box = await element.bounding_box()
                        if box:
                            center_x = box['x'] + box['width'] / 2
                            center_y = box['y'] + box['height'] / 2
                            await context.page.mouse.move(center_x, center_y)
                            await context.page.mouse.wheel(delta_x, delta_y)
                        else:
                            raise Exception("无法获取元素位置")
                    else:
                        # 在页面中心滚动
                        viewport = context.page.viewport_size
                        if viewport:
                            await context.page.mouse.move(viewport['width'] / 2, viewport['height'] / 2)
                        await context.page.mouse.wheel(delta_x, delta_y)
                    
                    return ModuleResult(success=True, message=f"已滚动 {direction} {distance}px (鼠标滚轮)")
                except Exception as wheel_error:
                    if scroll_mode == 'wheel':
                        raise wheel_error
                    # auto 模式下，滚轮失败则尝试脚本滚动
            
            # 使用 JavaScript 滚动
            if selector:
                await context.page.locator(format_selector(selector)).evaluate(
                    f"el => el.scrollBy({delta_x}, {delta_y})"
                )
            else:
                await context.page.evaluate(f"window.scrollBy({delta_x}, {delta_y})")
            
            return ModuleResult(success=True, message=f"已滚动 {direction} {distance}px")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"滚动页面失败: {str(e)}")


@register_executor
class UploadFileExecutor(ModuleExecutor):
    """上传文件模块执行器（已由 advanced_browser.UploadFileExecutor 覆盖，保持同步）"""

    @property
    def module_type(self) -> str:
        return "upload_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        # 此执行器会被 advanced_browser.UploadFileExecutor 覆盖（后注册优先）
        # 保持实现同步，以防注册顺序变化
        from .advanced_browser import UploadFileExecutor as _Impl
        return await _Impl().execute(config, context)


@register_executor
class DownloadFileExecutor(ModuleExecutor):
    """下载文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "download_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import httpx
        import os
        from urllib.parse import urlparse, unquote
        
        download_mode = context.resolve_value(config.get("downloadMode", "click"))  # 支持变量引用
        trigger_selector = context.resolve_value(config.get("triggerSelector", ""))
        download_url = context.resolve_value(config.get("downloadUrl", ""))
        save_path = context.resolve_value(config.get("savePath", ""))
        file_name = context.resolve_value(config.get("fileName", ""))
        variable_name = config.get("variableName", "")

        try:
            if download_mode == "url":
                if not download_url:
                    return ModuleResult(success=False, error="下载URL不能为空")

                if not file_name:
                    parsed = urlparse(download_url)
                    file_name = unquote(os.path.basename(parsed.path)) or "downloaded_file"

                if save_path:
                    Path(save_path).mkdir(parents=True, exist_ok=True)
                    final_path = Path(save_path) / file_name
                else:
                    downloads_dir = Path("downloads")
                    downloads_dir.mkdir(exist_ok=True)
                    final_path = downloads_dir / file_name

                async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
                    response = await client.get(download_url)
                    response.raise_for_status()

                    with open(final_path, "wb") as f:
                        f.write(response.content)

                if variable_name:
                    context.set_variable(variable_name, str(final_path))

                return ModuleResult(
                    success=True,
                    message=f"已下载文件: {final_path}",
                    data=str(final_path),
                )

            else:
                if not trigger_selector:
                    return ModuleResult(success=False, error="触发元素选择器不能为空")

                if context.page is None:
                    return ModuleResult(success=False, error="没有打开的页面")

                await context.switch_to_latest_page()

                async with context.page.expect_download() as download_info:
                    await context.page.click(trigger_selector)

                download = await download_info.value

                if save_path:
                    Path(save_path).mkdir(parents=True, exist_ok=True)
                    if file_name:
                        final_path = Path(save_path) / file_name
                    else:
                        final_path = Path(save_path) / download.suggested_filename
                    await download.save_as(str(final_path))
                else:
                    final_path = await download.path()

                if variable_name:
                    context.set_variable(variable_name, str(final_path))

                return ModuleResult(
                    success=True,
                    message=f"已下载文件: {final_path}",
                    data=str(final_path),
                )

        except Exception as e:
            return ModuleResult(success=False, error=f"下载文件失败: {str(e)}")


@register_executor
class SaveImageExecutor(ModuleExecutor):
    """保存图片模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "save_image"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get('selector', ''))
        save_path = context.resolve_value(config.get('savePath', ''))
        variable_name = config.get('variableName', '')
        
        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            
            element = context.page.locator(format_selector(selector))
            src = await element.get_attribute('src')
            
            if src and src.startswith('data:'):
                header, data = src.split(',', 1)
                image_data = base64.b64decode(data)
            else:
                image_data = await element.screenshot()
            
            if save_path:
                Path(save_path).parent.mkdir(parents=True, exist_ok=True)
                with open(save_path, 'wb') as f:
                    f.write(image_data)
                final_path = save_path
            else:
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
                    f.write(image_data)
                    final_path = f.name
            
            if variable_name:
                context.set_variable(variable_name, final_path)
            
            return ModuleResult(success=True, message=f"已保存图片: {final_path}", data=final_path)
        
        except Exception as e:
            return ModuleResult(success=False, error=f"保存图片失败: {str(e)}")


@register_executor
class SendEmailExecutor(ModuleExecutor):
    """发送邮件模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "send_email"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        sender_email = context.resolve_value(config.get('senderEmail', ''))
        auth_code = context.resolve_value(config.get('authCode', ''))
        recipient_email = context.resolve_value(config.get('recipientEmail', ''))
        email_subject = context.resolve_value(config.get('emailSubject', ''))
        email_content = context.resolve_value(config.get('emailContent', ''))
        
        if not sender_email:
            return ModuleResult(success=False, error="发件人邮箱不能为空")
        
        if not auth_code:
            return ModuleResult(success=False, error="授权码不能为空")
        
        if not recipient_email:
            return ModuleResult(success=False, error="收件人邮箱不能为空")
        
        try:
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = recipient_email
            msg['Subject'] = email_subject or '(无标题)'
            
            msg.attach(MIMEText(email_content or '', 'plain', 'utf-8'))
            
            # 使用线程池执行同步SMTP操作
            loop = asyncio.get_running_loop()
            await loop.run_in_executor(None, self._send_email_sync, sender_email, auth_code, recipient_email, msg)
            
            return ModuleResult(
                success=True, 
                message=f"邮件已发送至: {recipient_email}",
                data={'recipient': recipient_email, 'subject': email_subject}
            )
        
        except smtplib.SMTPAuthenticationError:
            return ModuleResult(success=False, error="邮箱认证失败，请检查邮箱地址和授权码")
        except Exception as e:
            return ModuleResult(success=False, error=f"发送邮件失败: {str(e)}")
    
    def _send_email_sync(self, sender_email, auth_code, recipient_email, msg):
        import smtplib
        server = smtplib.SMTP_SSL('smtp.qq.com', 465)
        server.login(sender_email, auth_code)
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()


@register_executor
class ReadExcelExecutor(ModuleExecutor):
    """Excel文件读取模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "read_excel"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import openpyxl
        from app.api.data_assets import get_asset_by_name
        
        file_name = context.resolve_value(config.get('fileName', ''))
        sheet_name = context.resolve_value(config.get('sheetName', ''))
        read_mode = context.resolve_value(config.get('readMode', 'cell'))  # 支持变量引用
        cell_address = context.resolve_value(config.get('cellAddress', ''))
        row_index = to_int(config.get('rowIndex', 1), 1, context)
        column_index = context.resolve_value(config.get('columnIndex', ''))
        start_cell = context.resolve_value(config.get('startCell', ''))
        end_cell = context.resolve_value(config.get('endCell', ''))
        variable_name = config.get('variableName', '')
        start_row = to_int(config.get('startRow', 2), 2, context)
        start_col = context.resolve_value(config.get('startCol', ''))
        
        if not file_name:
            return ModuleResult(success=False, error="请选择要读取的Excel文件")
        
        if not variable_name:
            return ModuleResult(success=False, error="请指定存储变量名")
        
        asset = get_asset_by_name(file_name)
        if not asset:
            return ModuleResult(success=False, error=f"文件 '{file_name}' 不存在")
        
        file_path = asset['path']
        is_xls = file_path.lower().endswith('.xls')
        
        try:
            # 使用线程池执行同步Excel操作
            loop = asyncio.get_running_loop()
            if is_xls:
                result, result_type = await loop.run_in_executor(
                    None, self._read_xls, file_path, sheet_name, read_mode, 
                    cell_address, row_index, column_index, start_cell, end_cell, start_row, start_col
                )
            else:
                result, result_type = await loop.run_in_executor(
                    None, self._read_xlsx, file_path, sheet_name, read_mode,
                    cell_address, row_index, column_index, start_cell, end_cell, start_row, start_col
                )
            
            context.set_variable(variable_name, result)
            
            display_content = str(result)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(
                success=True,
                message=f"已读取Excel数据: {display_content}",
                data={'value': result, 'type': result_type, 'file': file_name}
            )
        except Exception as e:
            return ModuleResult(success=False, error=f"读取Excel失败: {str(e)}")
    
    def _read_xlsx(self, file_path, sheet_name, read_mode, cell_address, row_index, 
                   column_index, start_cell, end_cell, start_row, start_col):
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise Exception(f"工作表 '{sheet_name}' 不存在")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        result = None
        result_type = 'unknown'
        
        if read_mode == 'cell':
            if not cell_address:
                wb.close()
                raise Exception("单元格模式需要指定单元格地址")
            cell = ws[cell_address]
            result = cell.value
            result_type = 'cell'
        
        elif read_mode == 'row':
            if row_index is None or row_index < 1:
                wb.close()
                raise Exception("行模式需要指定有效的行号")
            row_data = []
            start_col_idx = 1
            if start_col:
                if isinstance(start_col, str) and start_col.isalpha():
                    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
                else:
                    start_col_idx = int(start_col)
            for cell in ws[row_index]:
                if cell.column >= start_col_idx:
                    row_data.append(cell.value)
            result = row_data
            result_type = 'array'
        
        elif read_mode == 'column':
            if not column_index:
                wb.close()
                raise Exception("列模式需要指定列号或列字母")
            col_data = []
            col_idx = column_index
            if isinstance(col_idx, str) and col_idx.isalpha():
                col_idx = openpyxl.utils.column_index_from_string(col_idx)
            else:
                col_idx = int(col_idx)
            for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
                col_data.append(row[0].value)
            result = col_data
            result_type = 'array'
        
        elif read_mode == 'range':
            if not start_cell or not end_cell:
                wb.close()
                raise Exception("范围模式需要指定起始和结束单元格")
            range_data = []
            for row in ws[f"{start_cell}:{end_cell}"]:
                row_data = [cell.value for cell in row]
                range_data.append(row_data)
            result = range_data
            result_type = 'matrix'
        
        wb.close()
        return result, result_type
    
    def _read_xls(self, file_path, sheet_name, read_mode, cell_address, row_index,
                  column_index, start_cell, end_cell, start_row, start_col):
        import xlrd
        
        wb = xlrd.open_workbook(file_path)
        
        if sheet_name:
            if sheet_name not in wb.sheet_names():
                raise Exception(f"工作表 '{sheet_name}' 不存在")
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
        
        result = None
        result_type = 'unknown'
        
        if read_mode == 'cell':
            if not cell_address:
                raise Exception("单元格模式需要指定单元格地址")
            col_idx, row_idx = self._parse_cell_address(cell_address)
            result = ws.cell_value(row_idx, col_idx)
            result_type = 'cell'
        
        elif read_mode == 'row':
            if row_index is None or row_index < 1:
                raise Exception("行模式需要指定有效的行号")
            row_data = ws.row_values(row_index - 1)
            if start_col:
                start_col_idx = 0
                if isinstance(start_col, str) and start_col.isalpha():
                    start_col_idx = self._col_letter_to_index(start_col)
                else:
                    start_col_idx = int(start_col) - 1
                row_data = row_data[start_col_idx:]
            result = row_data
            result_type = 'array'
        
        elif read_mode == 'column':
            if not column_index:
                raise Exception("列模式需要指定列号或列字母")
            col_idx = column_index
            if isinstance(col_idx, str) and col_idx.isalpha():
                col_idx = self._col_letter_to_index(col_idx)
            else:
                col_idx = int(col_idx) - 1
            col_data = ws.col_values(col_idx, start_rowx=start_row - 1)
            result = col_data
            result_type = 'array'
        
        elif read_mode == 'range':
            if not start_cell or not end_cell:
                raise Exception("范围模式需要指定起始和结束单元格")
            start_col_idx, start_row_idx = self._parse_cell_address(start_cell)
            end_col_idx, end_row_idx = self._parse_cell_address(end_cell)
            range_data = []
            for r in range(start_row_idx, end_row_idx + 1):
                row_data = []
                for c in range(start_col_idx, end_col_idx + 1):
                    row_data.append(ws.cell_value(r, c))
                range_data.append(row_data)
            result = range_data
            result_type = 'matrix'
        
        return result, result_type
    
    def _col_letter_to_index(self, col_str: str) -> int:
        result = 0
        for c in col_str.upper():
            result = result * 26 + (ord(c) - ord('A') + 1)
        return result - 1
    
    def _parse_cell_address(self, address: str) -> tuple:
        col_str, row_str = '', ''
        for c in address:
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        return self._col_letter_to_index(col_str), int(row_str) - 1


@register_executor
class SetClipboardExecutor(ModuleExecutor):
    """设置剪贴板模块执行器"""

    @property
    def module_type(self) -> str:
        return "set_clipboard"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import subprocess
        import ctypes
        from ctypes import wintypes
        
        content_type = context.resolve_value(config.get("contentType", "text"))  # 支持变量引用
        text_content = context.resolve_value(config.get("textContent", ""))
        image_path = context.resolve_value(config.get("imagePath", ""))

        try:
            if content_type == "image":
                if not image_path:
                    return ModuleResult(success=False, error="图片路径不能为空")

                if not Path(image_path).exists():
                    return ModuleResult(success=False, error=f"图片文件不存在: {image_path}")

                try:
                    from PIL import Image
                    import tempfile

                    img = Image.open(image_path)
                    if img.mode == "RGBA":
                        img = img.convert("RGB")
                    
                    with tempfile.NamedTemporaryFile(suffix='.bmp', delete=False) as tmp:
                        tmp_path = tmp.name
                        img.save(tmp_path, "BMP")
                    
                    ps_script = f'''
Add-Type -AssemblyName System.Windows.Forms
$image = [System.Drawing.Image]::FromFile("{tmp_path.replace(chr(92), '/')}")
[System.Windows.Forms.Clipboard]::SetImage($image)
$image.Dispose()
'''
                    loop = asyncio.get_running_loop()
                    result = await loop.run_in_executor(
                        None, 
                        lambda: subprocess.run(["powershell", "-Command", ps_script], capture_output=True, text=True)
                    )
                    
                    try:
                        Path(tmp_path).unlink()
                    except:
                        pass
                    
                    if result.returncode != 0:
                        return ModuleResult(success=False, error=f"设置剪贴板失败: {result.stderr}")

                    return ModuleResult(success=True, message=f"已将图片复制到剪贴板: {image_path}")

                except ImportError:
                    return ModuleResult(success=False, error="需要安装 Pillow 库")

            else:
                if not text_content:
                    return ModuleResult(success=False, error="文本内容不能为空")

                # 使用 Windows API 直接操作剪贴板（更稳定）
                def set_clipboard_text(text: str) -> tuple[bool, str]:
                    """使用 Windows API 设置剪贴板文本，返回 (成功, 错误信息)"""
                    import ctypes
                    from ctypes import wintypes
                    
                    user32 = ctypes.windll.user32
                    kernel32 = ctypes.windll.kernel32
                    
                    # 使用 HANDLE 类型（64位兼容）
                    HANDLE = ctypes.c_void_p
                    LPVOID = ctypes.c_void_p
                    
                    # 设置正确的参数和返回类型
                    kernel32.GlobalAlloc.argtypes = [wintypes.UINT, ctypes.c_size_t]
                    kernel32.GlobalAlloc.restype = HANDLE
                    kernel32.GlobalLock.argtypes = [HANDLE]
                    kernel32.GlobalLock.restype = LPVOID
                    kernel32.GlobalUnlock.argtypes = [HANDLE]
                    kernel32.GlobalUnlock.restype = wintypes.BOOL
                    kernel32.GlobalFree.argtypes = [HANDLE]
                    kernel32.GlobalFree.restype = HANDLE
                    user32.SetClipboardData.argtypes = [wintypes.UINT, HANDLE]
                    user32.SetClipboardData.restype = HANDLE
                    
                    CF_UNICODETEXT = 13
                    GMEM_MOVEABLE = 0x0002
                    
                    # 尝试多次打开剪贴板（处理被占用的情况）
                    max_retries = 10
                    for i in range(max_retries):
                        if user32.OpenClipboard(0):
                            break
                        import time
                        time.sleep(0.1)
                    else:
                        return False, "无法打开剪贴板，可能被其他程序占用"
                    
                    try:
                        user32.EmptyClipboard()
                        
                        # 转换为 UTF-16 编码（包含终止符）
                        text_bytes = text.encode('utf-16-le') + b'\x00\x00'
                        
                        # 分配全局内存
                        h_mem = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(text_bytes))
                        if not h_mem:
                            return False, "内存分配失败"
                        
                        # 锁定内存并复制数据
                        p_mem = kernel32.GlobalLock(h_mem)
                        if not p_mem:
                            kernel32.GlobalFree(h_mem)
                            return False, "内存锁定失败"
                        
                        try:
                            ctypes.memmove(p_mem, text_bytes, len(text_bytes))
                        finally:
                            kernel32.GlobalUnlock(h_mem)
                        
                        # 设置剪贴板数据
                        if not user32.SetClipboardData(CF_UNICODETEXT, h_mem):
                            kernel32.GlobalFree(h_mem)
                            return False, "设置剪贴板数据失败"
                        
                        return True, ""
                    finally:
                        user32.CloseClipboard()
                
                loop = asyncio.get_running_loop()
                success, error_msg = await loop.run_in_executor(None, lambda: set_clipboard_text(text_content))
                
                if not success:
                    return ModuleResult(success=False, error=f"设置剪贴板失败: {error_msg}")

                display_text = text_content[:50] + "..." if len(text_content) > 50 else text_content
                return ModuleResult(success=True, message=f"已将文本复制到剪贴板: {display_text}")

        except Exception as e:
            return ModuleResult(success=False, error=f"设置剪贴板失败: {str(e)}")


@register_executor
class GetClipboardExecutor(ModuleExecutor):
    """获取剪贴板模块执行器"""

    @property
    def module_type(self) -> str:
        return "get_clipboard"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        variable_name = config.get("variableName", "")

        if not variable_name:
            return ModuleResult(success=False, error="存储变量名不能为空")

        try:
            def get_clipboard_text() -> tuple[str | None, str]:
                """使用 Windows API 获取剪贴板文本，返回 (内容, 错误信息)"""
                import ctypes
                from ctypes import wintypes
                
                user32 = ctypes.windll.user32
                kernel32 = ctypes.windll.kernel32
                
                # 使用 HANDLE 类型（64位兼容）
                HANDLE = ctypes.c_void_p
                LPVOID = ctypes.c_void_p
                
                # 设置正确的参数和返回类型
                kernel32.GlobalLock.argtypes = [HANDLE]
                kernel32.GlobalLock.restype = LPVOID
                kernel32.GlobalUnlock.argtypes = [HANDLE]
                kernel32.GlobalUnlock.restype = wintypes.BOOL
                user32.GetClipboardData.argtypes = [wintypes.UINT]
                user32.GetClipboardData.restype = HANDLE
                
                CF_UNICODETEXT = 13
                
                # 尝试多次打开剪贴板（处理被占用的情况）
                max_retries = 10
                for i in range(max_retries):
                    if user32.OpenClipboard(0):
                        break
                    import time
                    time.sleep(0.1)
                else:
                    return None, "无法打开剪贴板，可能被其他程序占用"
                
                try:
                    # 获取剪贴板数据
                    h_data = user32.GetClipboardData(CF_UNICODETEXT)
                    if not h_data:
                        return "", ""  # 剪贴板为空
                    
                    # 锁定内存并读取数据
                    p_data = kernel32.GlobalLock(h_data)
                    if not p_data:
                        return "", ""
                    
                    try:
                        # 读取 Unicode 字符串
                        text = ctypes.wstring_at(p_data)
                        return text, ""
                    finally:
                        kernel32.GlobalUnlock(h_data)
                finally:
                    user32.CloseClipboard()
            
            loop = asyncio.get_running_loop()
            clipboard_content, error_msg = await loop.run_in_executor(None, get_clipboard_text)
            
            if clipboard_content is None:
                return ModuleResult(success=False, error=f"获取剪贴板失败: {error_msg}")
            
            context.set_variable(variable_name, clipboard_content)

            if not clipboard_content:
                return ModuleResult(
                    success=True,
                    message=f"剪贴板为空，已设置变量 {variable_name} 为空字符串",
                    data=clipboard_content
                )

            display_text = clipboard_content[:50] + "..." if len(clipboard_content) > 50 else clipboard_content
            return ModuleResult(
                success=True,
                message=f"已获取剪贴板内容: {display_text}",
                data=clipboard_content
            )

        except Exception as e:
            return ModuleResult(success=False, error=f"获取剪贴板失败: {str(e)}")


@register_executor
class KeyboardActionExecutor(ModuleExecutor):
    """键盘操作模块执行器"""

    @property
    def module_type(self) -> str:
        return "keyboard_action"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        target_type = context.resolve_value(config.get("targetType", "page"))  # 支持变量引用
        selector = context.resolve_value(config.get("selector", ""))
        key_sequence = context.resolve_value(config.get("keySequence", ""))
        delay = to_int(config.get("delay", 0), 0, context)
        press_mode = context.resolve_value(config.get("pressMode", "click"))  # 支持变量引用
        hold_duration = to_float(config.get("holdDuration", 1), 1, context)  # 长按时长(秒)

        if not key_sequence:
            return ModuleResult(success=False, error="按键序列不能为空")

        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")

        try:
            await context.switch_to_latest_page()

            if target_type == "element" and selector:
                element = context.page.locator(format_selector(selector))
                await element.focus()
                await asyncio.sleep(0.1)

            # 键名映射：将用户友好的键名转换为 Playwright 需要的键名
            KEY_MAP = {
                'ctrl': 'Control',
                'Ctrl': 'Control',
                'CTRL': 'Control',
                'alt': 'Alt',
                'ALT': 'Alt',
                'shift': 'Shift',
                'SHIFT': 'Shift',
                'meta': 'Meta',
                'win': 'Meta',
                'Win': 'Meta',
                'WIN': 'Meta',
                'enter': 'Enter',
                'ENTER': 'Enter',
                'tab': 'Tab',
                'TAB': 'Tab',
                'esc': 'Escape',
                'ESC': 'Escape',
                'escape': 'Escape',
                'backspace': 'Backspace',
                'BACKSPACE': 'Backspace',
                'delete': 'Delete',
                'DELETE': 'Delete',
                'space': 'Space',
                'SPACE': 'Space',
                'up': 'ArrowUp',
                'UP': 'ArrowUp',
                'down': 'ArrowDown',
                'DOWN': 'ArrowDown',
                'left': 'ArrowLeft',
                'LEFT': 'ArrowLeft',
                'right': 'ArrowRight',
                'RIGHT': 'ArrowRight',
            }
            
            keys = key_sequence.split("+")
            # 转换键名
            keys = [KEY_MAP.get(k.strip(), k.strip()) for k in keys]

            if press_mode == "hold":
                # 长按模式：按下所有键，等待指定时间，然后释放
                for key in keys:
                    await context.page.keyboard.down(key)
                
                await asyncio.sleep(hold_duration)

                for key in reversed(keys):
                    await context.page.keyboard.up(key)

                return ModuleResult(success=True, message=f"已长按 {key_sequence} {hold_duration}秒")
            else:
                # 点击模式（默认）
                if len(keys) == 1:
                    await context.page.keyboard.press(keys[0])
                else:
                    modifiers = keys[:-1]
                    main_key = keys[-1]

                    for mod in modifiers:
                        await context.page.keyboard.down(mod)

                    if delay > 0:
                        await asyncio.sleep(delay / 1000)

                    await context.page.keyboard.press(main_key)

                    for mod in reversed(modifiers):
                        await context.page.keyboard.up(mod)

                return ModuleResult(success=True, message=f"已执行键盘操作: {key_sequence}")

        except Exception as e:
            return ModuleResult(success=False, error=f"键盘操作失败: {str(e)}")


@register_executor
class RealMouseScrollExecutor(ModuleExecutor):
    """真实鼠标滚动模块执行器 - 使用 SendInput API 实现真正的硬件级滚轮模拟"""

    @property
    def module_type(self) -> str:
        return "real_mouse_scroll"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        direction = context.resolve_value(config.get("direction", "down"))  # 支持变量引用
        scroll_amount = to_int(config.get("scrollAmount", 3), 3, context)  # 滚动格数
        scroll_count = to_int(config.get("scrollCount", 1), 1, context)  # 滚动次数
        scroll_interval = to_float(config.get("scrollInterval", 0.1), 0.1, context)  # 滚动间隔(秒)

        try:
            import ctypes
            from ctypes import wintypes
            
            # SendInput 结构体定义
            INPUT_MOUSE = 0
            MOUSEEVENTF_WHEEL = 0x0800
            WHEEL_DELTA = 120  # 一格滚轮的标准值
            
            class MOUSEINPUT(ctypes.Structure):
                _fields_ = [
                    ("dx", wintypes.LONG),
                    ("dy", wintypes.LONG),
                    ("mouseData", wintypes.DWORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class INPUT(ctypes.Structure):
                _fields_ = [
                    ("type", wintypes.DWORD),
                    ("mi", MOUSEINPUT)
                ]
            
            # 定义SendInput函数的参数类型（确保类型正确）
            user32 = ctypes.windll.user32
            user32.SendInput.argtypes = [
                wintypes.UINT,  # nInputs
                ctypes.POINTER(INPUT),  # pInputs
                ctypes.c_int  # cbSize
            ]
            user32.SendInput.restype = wintypes.UINT
            
            # 计算滚动量（向上为正，向下为负）
            delta = WHEEL_DELTA * scroll_amount
            if direction == "down":
                delta = -delta
            
            # 执行滚动
            for i in range(scroll_count):
                # 构建 INPUT 结构
                inp = INPUT()
                inp.type = INPUT_MOUSE
                inp.mi.dx = 0
                inp.mi.dy = 0
                inp.mi.mouseData = delta & 0xFFFFFFFF  # 转为无符号
                inp.mi.dwFlags = MOUSEEVENTF_WHEEL
                inp.mi.time = 0
                inp.mi.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                
                # 使用 SendInput 发送滚轮事件
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                
                if i < scroll_count - 1 and scroll_interval > 0:
                    await asyncio.sleep(scroll_interval / 1000)
            
            direction_text = "向下" if direction == "down" else "向上"
            return ModuleResult(
                success=True, 
                message=f"已{direction_text}滚动 {scroll_count} 次，每次 {scroll_amount} 格"
            )

        except ImportError:
            return ModuleResult(success=False, error="此功能仅支持 Windows 系统")
        except Exception as e:
            return ModuleResult(success=False, error=f"真实鼠标滚动失败: {str(e)}")


@register_executor
class ShutdownSystemExecutor(ModuleExecutor):
    """关机/重启模块执行器"""

    @property
    def module_type(self) -> str:
        return "shutdown_system"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import subprocess
        
        action = context.resolve_value(config.get("action", "shutdown"))  # 支持变量引用
        delay = to_int(config.get("delay", 0), 0, context)
        force_raw = config.get("force", False)
        # 支持变量引用
        if isinstance(force_raw, str):
            force_raw = context.resolve_value(force_raw)
        force = force_raw in [True, 'true', 'True', '1', 1]

        try:
            # 构建命令
            if action == "shutdown":
                cmd = f"shutdown /s /t {delay}"
                if force:
                    cmd += " /f"
                action_text = "关机"
            elif action == "restart":
                cmd = f"shutdown /r /t {delay}"
                if force:
                    cmd += " /f"
                action_text = "重启"
            elif action == "logout":
                cmd = "shutdown /l"
                action_text = "注销"
            elif action == "hibernate":
                cmd = "shutdown /h"
                action_text = "休眠"
            elif action == "sleep":
                # 使用 PowerShell 进入睡眠模式
                cmd = "powershell -Command \"Add-Type -AssemblyName System.Windows.Forms; [System.Windows.Forms.Application]::SetSuspendState('Suspend', $false, $false)\""
                action_text = "睡眠"
            else:
                return ModuleResult(success=False, error=f"未知操作类型: {action}")

            loop = asyncio.get_running_loop()
            result = await loop.run_in_executor(
                None,
                lambda: subprocess.run(cmd, shell=True, capture_output=True, text=True)
            )

            if result.returncode != 0 and result.stderr:
                return ModuleResult(success=False, error=f"执行失败: {result.stderr}")

            if delay > 0:
                return ModuleResult(success=True, message=f"系统将在 {delay} 秒后{action_text}")
            else:
                return ModuleResult(success=True, message=f"正在执行{action_text}...")

        except Exception as e:
            return ModuleResult(success=False, error=f"系统操作失败: {str(e)}")


@register_executor
class LockScreenExecutor(ModuleExecutor):
    """锁定屏幕模块执行器"""

    @property
    def module_type(self) -> str:
        return "lock_screen"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes

        try:
            # 调用 Windows API 锁定屏幕
            ctypes.windll.user32.LockWorkStation()
            return ModuleResult(success=True, message="屏幕已锁定")
        except Exception as e:
            return ModuleResult(success=False, error=f"锁定屏幕失败: {str(e)}")


@register_executor
class WindowFocusExecutor(ModuleExecutor):
    """窗口聚焦模块执行器 - 将指定窗口置顶到最前面"""

    @property
    def module_type(self) -> str:
        return "window_focus"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes
        from ctypes import wintypes
        import time
        
        window_title = context.resolve_value(config.get("windowTitle", ""))
        match_mode = context.resolve_value(config.get("matchMode", "contains"))  # exact, contains, startswith
        
        if not window_title:
            return ModuleResult(success=False, error="窗口标题不能为空")
        
        try:
            user32 = ctypes.windll.user32
            
            # 定义回调函数类型
            EnumWindowsProc = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
            
            found_hwnd = None
            found_title = None
            
            def enum_windows_callback(hwnd, lParam):
                nonlocal found_hwnd, found_title
                
                # 获取窗口标题
                length = user32.GetWindowTextLengthW(hwnd)
                if length == 0:
                    return True
                
                buffer = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(hwnd, buffer, length + 1)
                title = buffer.value
                
                # 检查窗口是否可见
                if not user32.IsWindowVisible(hwnd):
                    return True
                
                # 根据匹配模式检查标题
                matched = False
                if match_mode == "exact":
                    matched = title == window_title
                elif match_mode == "contains":
                    matched = window_title.lower() in title.lower()
                elif match_mode == "startswith":
                    matched = title.lower().startswith(window_title.lower())
                
                if matched:
                    found_hwnd = hwnd
                    found_title = title
                    return False  # 停止枚举
                
                return True
            
            # 枚举所有窗口
            callback = EnumWindowsProc(enum_windows_callback)
            user32.EnumWindows(callback, 0)
            
            if not found_hwnd:
                return ModuleResult(success=False, error=f"未找到匹配的窗口: {window_title}")
            
            # 定义常量
            SW_RESTORE = 9
            SW_SHOW = 5
            SW_SHOWNORMAL = 1
            HWND_TOPMOST = -1
            HWND_NOTOPMOST = -2
            SWP_NOMOVE = 0x0002
            SWP_NOSIZE = 0x0001
            SWP_SHOWWINDOW = 0x0040
            SWP_NOACTIVATE = 0x0010
            
            # 如果窗口最小化，先恢复
            if user32.IsIconic(found_hwnd):
                user32.ShowWindow(found_hwnd, SW_RESTORE)
                time.sleep(0.1)
            
            # 方法1: 使用 AttachThreadInput 绕过前台窗口限制
            # 获取目标窗口的线程ID
            target_thread_id = user32.GetWindowThreadProcessId(found_hwnd, None)
            # 获取当前线程ID
            current_thread_id = ctypes.windll.kernel32.GetCurrentThreadId()
            
            # 附加到目标线程的输入队列
            attached = False
            if target_thread_id != current_thread_id:
                attached = user32.AttachThreadInput(current_thread_id, target_thread_id, True)
            
            try:
                # 先将窗口设为最顶层
                user32.SetWindowPos(found_hwnd, HWND_TOPMOST, 0, 0, 0, 0, 
                                   SWP_NOMOVE | SWP_NOSIZE | SWP_SHOWWINDOW)
                
                # 显示窗口
                user32.ShowWindow(found_hwnd, SW_SHOW)
                
                # 激活窗口
                user32.BringWindowToTop(found_hwnd)
                user32.SetForegroundWindow(found_hwnd)
                
                # 模拟 Alt 键按下释放，这可以帮助绕过 Windows 的前台窗口限制
                # VK_MENU = 0x12 (Alt键)
                user32.keybd_event(0x12, 0, 0, 0)  # Alt down
                user32.keybd_event(0x12, 0, 2, 0)  # Alt up
                
                # 再次尝试设置前台窗口
                user32.SetForegroundWindow(found_hwnd)
                user32.SetActiveWindow(found_hwnd)
                
                # 取消最顶层（这样窗口不会一直置顶）
                user32.SetWindowPos(found_hwnd, HWND_NOTOPMOST, 0, 0, 0, 0, 
                                   SWP_NOMOVE | SWP_NOSIZE | SWP_SHOWWINDOW)
                
            finally:
                # 分离线程输入
                if attached:
                    user32.AttachThreadInput(current_thread_id, target_thread_id, False)
            
            return ModuleResult(
                success=True, 
                message=f"已聚焦窗口: {found_title}",
                data={'hwnd': found_hwnd, 'title': found_title}
            )
            
        except Exception as e:
            return ModuleResult(success=False, error=f"窗口聚焦失败: {str(e)}")


@register_executor
class RealMouseClickExecutor(ModuleExecutor):
    """真实鼠标点击模块执行器 - 使用 SendInput API 实现真正的硬件级鼠标点击"""

    @property
    def module_type(self) -> str:
        return "real_mouse_click"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes
        from ctypes import wintypes

        x = context.resolve_value(config.get("x", ""))
        y = context.resolve_value(config.get("y", ""))
        button = context.resolve_value(config.get("button", "left"))  # 支持变量引用
        click_type = context.resolve_value(config.get("clickType", "single"))  # single/double/hold
        hold_duration = to_float(config.get("holdDuration", 1), 1, context)  # 长按时长(秒)

        if not x or not y:
            return ModuleResult(success=False, error="X和Y坐标不能为空")

        try:
            x = int(x)
            y = int(y)
        except ValueError:
            return ModuleResult(success=False, error="坐标必须是数字")

        try:
            # 设置 DPI 感知，确保坐标准确
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except:
                    pass
            
            user32 = ctypes.windll.user32
            
            # SendInput 结构体定义
            INPUT_MOUSE = 0
            MOUSEEVENTF_LEFTDOWN = 0x0002
            MOUSEEVENTF_LEFTUP = 0x0004
            MOUSEEVENTF_RIGHTDOWN = 0x0008
            MOUSEEVENTF_RIGHTUP = 0x0010
            MOUSEEVENTF_MIDDLEDOWN = 0x0020
            MOUSEEVENTF_MIDDLEUP = 0x0040
            
            class MOUSEINPUT(ctypes.Structure):
                _fields_ = [
                    ("dx", wintypes.LONG),
                    ("dy", wintypes.LONG),
                    ("mouseData", wintypes.DWORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class INPUT(ctypes.Structure):
                _fields_ = [
                    ("type", wintypes.DWORD),
                    ("mi", MOUSEINPUT)
                ]

            # 根据按键类型选择事件
            if button == "left":
                down_event = MOUSEEVENTF_LEFTDOWN
                up_event = MOUSEEVENTF_LEFTUP
            elif button == "right":
                down_event = MOUSEEVENTF_RIGHTDOWN
                up_event = MOUSEEVENTF_RIGHTUP
            else:  # middle
                down_event = MOUSEEVENTF_MIDDLEDOWN
                up_event = MOUSEEVENTF_MIDDLEUP

            button_text = {"left": "左键", "right": "右键", "middle": "中键"}[button]
            
            # 定义SendInput函数的参数类型（确保类型正确）
            user32.SendInput.argtypes = [
                wintypes.UINT,  # nInputs
                ctypes.POINTER(INPUT),  # pInputs
                ctypes.c_int  # cbSize
            ]
            user32.SendInput.restype = wintypes.UINT
            
            # 使用 SetCursorPos 直接移动鼠标（不需要坐标转换，更精确）
            def move_mouse(px, py):
                user32.SetCursorPos(int(px), int(py))
            
            # 发送鼠标按键事件
            def send_mouse_event(event_flag):
                inp = INPUT()
                inp.type = INPUT_MOUSE
                inp.mi.dx = 0
                inp.mi.dy = 0
                inp.mi.mouseData = 0
                inp.mi.dwFlags = event_flag
                inp.mi.time = 0
                inp.mi.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                # 使用ctypes.pointer传递结构体指针
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result

            if click_type == "hold":
                # 长按模式
                move_mouse(x, y)
                await asyncio.sleep(0.02)
                send_mouse_event(down_event)
                await asyncio.sleep(hold_duration)
                send_mouse_event(up_event)

                return ModuleResult(
                    success=True,
                    message=f"已在 ({x}, {y}) 执行{button_text}长按 {hold_duration}秒"
                )
            else:
                # 单击或双击模式
                click_count = 2 if click_type == "double" else 1
                move_mouse(x, y)
                await asyncio.sleep(0.02)
                
                for _ in range(click_count):
                    send_mouse_event(down_event)
                    await asyncio.sleep(0.05)
                    send_mouse_event(up_event)
                    if click_type == "double":
                        await asyncio.sleep(0.1)

                click_text = "双击" if click_type == "double" else "单击"
                return ModuleResult(
                    success=True, 
                    message=f"已在 ({x}, {y}) 执行{button_text}{click_text}"
                )

        except Exception as e:
            return ModuleResult(success=False, error=f"鼠标点击失败: {str(e)}")


@register_executor
class RealMouseMoveExecutor(ModuleExecutor):
    """真实鼠标移动模块执行器 - 使用 SetCursorPos 实现精确的鼠标移动"""

    @property
    def module_type(self) -> str:
        return "real_mouse_move"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes

        x = context.resolve_value(config.get("x", ""))
        y = context.resolve_value(config.get("y", ""))
        duration = to_int(config.get("duration", 0), 0, context)

        if not x or not y:
            return ModuleResult(success=False, error="X和Y坐标不能为空")

        try:
            target_x = int(x)
            target_y = int(y)
        except ValueError:
            return ModuleResult(success=False, error="坐标必须是数字")

        try:
            # 设置 DPI 感知，确保坐标准确
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except:
                    pass
            
            user32 = ctypes.windll.user32

            if duration > 0:
                # 平滑移动
                class POINT(ctypes.Structure):
                    _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]
                
                pt = POINT()
                user32.GetCursorPos(ctypes.byref(pt))
                start_x, start_y = pt.x, pt.y

                steps = max(10, duration // 10)
                for i in range(steps + 1):
                    progress = i / steps
                    current_x = int(start_x + (target_x - start_x) * progress)
                    current_y = int(start_y + (target_y - start_y) * progress)
                    user32.SetCursorPos(current_x, current_y)
                    await asyncio.sleep(duration / 1000 / steps)
            else:
                # 瞬间移动
                user32.SetCursorPos(target_x, target_y)

            return ModuleResult(
                success=True, 
                message=f"鼠标已移动到 ({target_x}, {target_y})"
            )

        except Exception as e:
            return ModuleResult(success=False, error=f"鼠标移动失败: {str(e)}")


@register_executor
class RealMouseDragExecutor(ModuleExecutor):
    """真实鼠标拖拽模块执行器 - 使用 SetCursorPos + SendInput 实现精确的鼠标拖拽"""

    @property
    def module_type(self) -> str:
        return "real_mouse_drag"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes
        from ctypes import wintypes

        # 获取起点和终点坐标
        start_x = context.resolve_value(config.get("startX", ""))
        start_y = context.resolve_value(config.get("startY", ""))
        end_x = context.resolve_value(config.get("endX", ""))
        end_y = context.resolve_value(config.get("endY", ""))
        button = context.resolve_value(config.get("button", "left"))  # 支持变量引用
        duration = to_float(config.get("duration", 0.5), 0.5, context)  # 拖拽时长，默认0.5秒

        # 验证坐标
        if not start_x or not start_y:
            return ModuleResult(success=False, error="起点坐标不能为空")
        if not end_x or not end_y:
            return ModuleResult(success=False, error="终点坐标不能为空")

        try:
            start_x = int(start_x)
            start_y = int(start_y)
            end_x = int(end_x)
            end_y = int(end_y)
        except ValueError:
            return ModuleResult(success=False, error="坐标必须是数字")

        try:
            # 设置 DPI 感知，确保坐标准确
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except:
                    pass
            
            user32 = ctypes.windll.user32
            
            # SendInput 结构体定义
            INPUT_MOUSE = 0
            MOUSEEVENTF_LEFTDOWN = 0x0002
            MOUSEEVENTF_LEFTUP = 0x0004
            MOUSEEVENTF_RIGHTDOWN = 0x0008
            MOUSEEVENTF_RIGHTUP = 0x0010
            MOUSEEVENTF_MIDDLEDOWN = 0x0020
            MOUSEEVENTF_MIDDLEUP = 0x0040
            
            class MOUSEINPUT(ctypes.Structure):
                _fields_ = [
                    ("dx", wintypes.LONG),
                    ("dy", wintypes.LONG),
                    ("mouseData", wintypes.DWORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class INPUT(ctypes.Structure):
                _fields_ = [
                    ("type", wintypes.DWORD),
                    ("mi", MOUSEINPUT)
                ]

            # 根据按键类型选择事件
            if button == "left":
                down_event = MOUSEEVENTF_LEFTDOWN
                up_event = MOUSEEVENTF_LEFTUP
            elif button == "right":
                down_event = MOUSEEVENTF_RIGHTDOWN
                up_event = MOUSEEVENTF_RIGHTUP
            else:  # middle
                down_event = MOUSEEVENTF_MIDDLEDOWN
                up_event = MOUSEEVENTF_MIDDLEUP
            
            # 定义SendInput函数的参数类型（确保类型正确）
            user32.SendInput.argtypes = [
                wintypes.UINT,  # nInputs
                ctypes.POINTER(INPUT),  # pInputs
                ctypes.c_int  # cbSize
            ]
            user32.SendInput.restype = wintypes.UINT
            
            # 发送鼠标按键事件
            def send_mouse_event(event_flag):
                inp = INPUT()
                inp.type = INPUT_MOUSE
                inp.mi.dx = 0
                inp.mi.dy = 0
                inp.mi.mouseData = 0
                inp.mi.dwFlags = event_flag
                inp.mi.time = 0
                inp.mi.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                # 使用ctypes.pointer传递结构体指针
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result

            # 1. 移动到起点
            user32.SetCursorPos(start_x, start_y)
            await asyncio.sleep(0.05)
            
            # 2. 按下鼠标
            send_mouse_event(down_event)
            await asyncio.sleep(0.05)

            # 3. 平滑拖拽到终点
            steps = max(10, int(duration * 1000) // 20)  # 至少10步，每步约20ms
            for i in range(1, steps + 1):
                progress = i / steps
                current_x = int(start_x + (end_x - start_x) * progress)
                current_y = int(start_y + (end_y - start_y) * progress)
                user32.SetCursorPos(current_x, current_y)
                await asyncio.sleep(duration / steps)

            # 4. 释放鼠标
            await asyncio.sleep(0.05)
            send_mouse_event(up_event)

            button_text = {"left": "左键", "right": "右键", "middle": "中键"}[button]
            return ModuleResult(
                success=True, 
                message=f"已使用{button_text}从 ({start_x}, {start_y}) 拖拽到 ({end_x}, {end_y})"
            )

        except Exception as e:
            return ModuleResult(success=False, error=f"鼠标拖拽失败: {str(e)}")


@register_executor
class RealKeyboardExecutor(ModuleExecutor):
    """真实键盘操作模块执行器 - 使用 SendInput API 实现真正的硬件级键盘输入"""

    @property
    def module_type(self) -> str:
        return "real_keyboard"

    # 虚拟键码映射
    VK_CODES = {
        'enter': 0x0D, 'tab': 0x09, 'escape': 0x1B, 'backspace': 0x08,
        'delete': 0x2E, 'space': 0x20, 'up': 0x26, 'down': 0x28,
        'left': 0x25, 'right': 0x27, 'home': 0x24, 'end': 0x23,
        'pageup': 0x21, 'pagedown': 0x22,
        'f1': 0x70, 'f2': 0x71, 'f3': 0x72, 'f4': 0x73, 'f5': 0x74,
        'f6': 0x75, 'f7': 0x76, 'f8': 0x77, 'f9': 0x78, 'f10': 0x79,
        'f11': 0x7A, 'f12': 0x7B,
        'ctrl': 0x11, 'alt': 0x12, 'shift': 0x10, 'win': 0x5B,
        'a': 0x41, 'b': 0x42, 'c': 0x43, 'd': 0x44, 'e': 0x45,
        'f': 0x46, 'g': 0x47, 'h': 0x48, 'i': 0x49, 'j': 0x4A,
        'k': 0x4B, 'l': 0x4C, 'm': 0x4D, 'n': 0x4E, 'o': 0x4F,
        'p': 0x50, 'q': 0x51, 'r': 0x52, 's': 0x53, 't': 0x54,
        'u': 0x55, 'v': 0x56, 'w': 0x57, 'x': 0x58, 'y': 0x59, 'z': 0x5A,
        '0': 0x30, '1': 0x31, '2': 0x32, '3': 0x33, '4': 0x34,
        '5': 0x35, '6': 0x36, '7': 0x37, '8': 0x38, '9': 0x39,
    }

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes
        from ctypes import wintypes

        input_type = context.resolve_value(config.get("inputType", "text"))  # 支持变量引用
        press_mode = context.resolve_value(config.get("pressMode", "click"))  # 支持变量引用
        hold_duration = to_float(config.get("holdDuration", 1), 1, context)  # 长按时长(秒)
        
        try:
            # SendInput 结构体定义
            INPUT_KEYBOARD = 1
            KEYEVENTF_KEYUP = 0x0002
            KEYEVENTF_UNICODE = 0x0004
            KEYEVENTF_SCANCODE = 0x0008
            
            class KEYBDINPUT(ctypes.Structure):
                _fields_ = [
                    ("wVk", wintypes.WORD),
                    ("wScan", wintypes.WORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class INPUT_UNION(ctypes.Union):
                _fields_ = [("ki", KEYBDINPUT)]
            
            class INPUT(ctypes.Structure):
                _fields_ = [
                    ("type", wintypes.DWORD),
                    ("union", INPUT_UNION)
                ]
            
            # 定义SendInput函数的参数类型（确保类型正确）
            user32 = ctypes.windll.user32
            user32.SendInput.argtypes = [
                wintypes.UINT,  # nInputs
                ctypes.POINTER(INPUT),  # pInputs
                ctypes.c_int  # cbSize
            ]
            user32.SendInput.restype = wintypes.UINT
            
            def send_key_down(vk_code):
                """发送按键按下事件"""
                inp = INPUT()
                inp.type = INPUT_KEYBOARD
                inp.union.ki.wVk = vk_code
                inp.union.ki.wScan = user32.MapVirtualKeyW(vk_code, 0)
                inp.union.ki.dwFlags = 0
                inp.union.ki.time = 0
                inp.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result
            
            def send_key_up(vk_code):
                """发送按键释放事件"""
                inp = INPUT()
                inp.type = INPUT_KEYBOARD
                inp.union.ki.wVk = vk_code
                inp.union.ki.wScan = user32.MapVirtualKeyW(vk_code, 0)
                inp.union.ki.dwFlags = KEYEVENTF_KEYUP
                inp.union.ki.time = 0
                inp.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result
            
            def send_unicode_char(char):
                """发送 Unicode 字符"""
                # 按下
                inp_down = INPUT()
                inp_down.type = INPUT_KEYBOARD
                inp_down.union.ki.wVk = 0
                inp_down.union.ki.wScan = ord(char)
                inp_down.union.ki.dwFlags = KEYEVENTF_UNICODE
                inp_down.union.ki.time = 0
                inp_down.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp_down), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                
                # 释放
                inp_up = INPUT()
                inp_up.type = INPUT_KEYBOARD
                inp_up.union.ki.wVk = 0
                inp_up.union.ki.wScan = ord(char)
                inp_up.union.ki.dwFlags = KEYEVENTF_UNICODE | KEYEVENTF_KEYUP
                inp_up.union.ki.time = 0
                inp_up.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp_up), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result

            if input_type == "text":
                text = context.resolve_value(config.get("text", ""))
                interval = to_float(config.get("interval", 0.05), 0.05, context)  # 每字符间隔(秒)

                if not text:
                    return ModuleResult(success=False, error="输入文本不能为空")

                # 使用 SendInput Unicode 输入
                for char in text:
                    send_unicode_char(char)
                    await asyncio.sleep(interval)

                display_text = text[:30] + "..." if len(text) > 30 else text
                return ModuleResult(success=True, message=f"已输入文本: {display_text}")

            elif input_type == "key":
                key = context.resolve_value(config.get("key", "enter")).lower()  # 支持变量引用
                vk_code = self.VK_CODES.get(key)
                
                if vk_code is None:
                    return ModuleResult(success=False, error=f"不支持的按键: {key}")

                if press_mode == "hold":
                    # 长按模式
                    send_key_down(vk_code)
                    await asyncio.sleep(hold_duration)
                    send_key_up(vk_code)
                    return ModuleResult(success=True, message=f"已长按按键: {key.upper()} {hold_duration}秒")
                else:
                    # 点击模式
                    send_key_down(vk_code)
                    await asyncio.sleep(0.05)
                    send_key_up(vk_code)
                    return ModuleResult(success=True, message=f"已按下按键: {key.upper()}")

            elif input_type == "hotkey":
                hotkey = context.resolve_value(config.get("hotkey", "")).lower()  # 支持变量引用
                
                if not hotkey:
                    return ModuleResult(success=False, error="组合键不能为空")

                keys = [k.strip() for k in hotkey.split("+")]
                vk_codes = []
                
                for key in keys:
                    vk_code = self.VK_CODES.get(key)
                    if vk_code is None:
                        return ModuleResult(success=False, error=f"不支持的按键: {key}")
                    vk_codes.append(vk_code)

                # 按下所有键
                for vk_code in vk_codes:
                    send_key_down(vk_code)
                    await asyncio.sleep(0.02)

                if press_mode == "hold":
                    # 长按模式
                    await asyncio.sleep(hold_duration)
                else:
                    await asyncio.sleep(0.05)

                # 释放所有键（逆序）
                for vk_code in reversed(vk_codes):
                    send_key_up(vk_code)
                    await asyncio.sleep(0.02)

                if press_mode == "hold":
                    return ModuleResult(success=True, message=f"已长按组合键: {hotkey.upper()} {hold_duration}秒")
                else:
                    return ModuleResult(success=True, message=f"已执行组合键: {hotkey.upper()}")

            else:
                return ModuleResult(success=False, error=f"未知输入类型: {input_type}")

        except Exception as e:
            return ModuleResult(success=False, error=f"键盘操作失败: {str(e)}")


@register_executor
class RunCommandExecutor(ModuleExecutor):
    """执行命令模块执行器"""

    @property
    def module_type(self) -> str:
        return "run_command"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import subprocess

        command = context.resolve_value(config.get("command", ""))
        shell = context.resolve_value(config.get("shell", "cmd"))  # 支持变量引用
        timeout = to_int(config.get("timeout", 30), 30, context)
        variable_name = config.get("variableName", "")

        if not command:
            return ModuleResult(success=False, error="命令不能为空")

        try:
            if shell == "powershell":
                full_command = ["powershell", "-Command", command]
            else:
                full_command = ["cmd", "/c", command]

            loop = asyncio.get_running_loop()
            result = await loop.run_in_executor(
                None,
                lambda: subprocess.run(
                    full_command,
                    capture_output=True,
                    timeout=timeout
                )
            )

            # 尝试解码输出
            try:
                stdout = result.stdout.decode('utf-8')
            except UnicodeDecodeError:
                stdout = result.stdout.decode('gbk', errors='ignore')

            try:
                stderr = result.stderr.decode('utf-8')
            except UnicodeDecodeError:
                stderr = result.stderr.decode('gbk', errors='ignore')

            output = stdout.strip() if stdout else stderr.strip()

            if variable_name:
                context.set_variable(variable_name, output)

            if result.returncode != 0:
                return ModuleResult(
                    success=False, 
                    error=f"命令执行失败 (返回码: {result.returncode}): {stderr or stdout}"
                )

            display_output = output[:100] + "..." if len(output) > 100 else output
            return ModuleResult(
                success=True, 
                message=f"命令执行成功: {display_output}",
                data={"output": output, "return_code": result.returncode}
            )

        except subprocess.TimeoutExpired:
            return ModuleResult(success=False, error=f"命令执行超时 ({timeout}秒)")
        except Exception as e:
            return ModuleResult(success=False, error=f"命令执行失败: {str(e)}")


@register_executor
class ClickImageExecutor(ModuleExecutor):
    """点击图像模块执行器 - 在屏幕上查找指定图像并点击"""

    @property
    def module_type(self) -> str:
        return "click_image"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes

        image_path = context.resolve_value(config.get("imagePath", ""))
        confidence = to_float(config.get("confidence", 0.8), 0.8, context)  # 匹配置信度
        button = context.resolve_value(config.get("button", "left"))  # 鼠标按键
        click_type = context.resolve_value(config.get("clickType", "single"))  # 点击类型
        click_position = context.resolve_value(config.get("clickPosition", "center"))  # 点击位置
        wait_timeout = to_int(config.get("waitTimeout", 10), 10, context)  # 等待超时
        search_region = config.get("searchRegion", None)  # 搜索区域

        if not image_path:
            return ModuleResult(success=False, error="图像路径不能为空")

        if not Path(image_path).exists():
            return ModuleResult(success=False, error=f"图像文件不存在: {image_path}")

        try:
            import cv2
            import numpy as np
            
            # 设置 DPI 感知，确保坐标准确
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except:
                    pass
            
            # 读取模板图像（支持中文路径）
            template = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_COLOR)
            if template is None:
                return ModuleResult(success=False, error="无法读取图像文件，请检查图像格式")

            template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
            h, w = template_gray.shape

            start_time = time.time()
            found = False
            click_x, click_y = 0, 0
            best_confidence = 0

            # 解析搜索区域（支持两点模式和起点+宽高模式）
            region_x, region_y, region_w, region_h = parse_search_region(search_region)
            use_region = region_w > 0 and region_h > 0

            # 获取完整的虚拟屏幕尺寸（支持多显示器）
            SM_XVIRTUALSCREEN = 76
            SM_YVIRTUALSCREEN = 77
            SM_CXVIRTUALSCREEN = 78
            SM_CYVIRTUALSCREEN = 79
            
            virtual_left = ctypes.windll.user32.GetSystemMetrics(SM_XVIRTUALSCREEN)
            virtual_top = ctypes.windll.user32.GetSystemMetrics(SM_YVIRTUALSCREEN)
            virtual_width = ctypes.windll.user32.GetSystemMetrics(SM_CXVIRTUALSCREEN)
            virtual_height = ctypes.windll.user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)

            while time.time() - start_time < wait_timeout:
                # 截取屏幕
                if use_region:
                    # 使用指定区域截图
                    from PIL import ImageGrab
                    screenshot_pil = ImageGrab.grab(bbox=(region_x, region_y, region_x + region_w, region_y + region_h))
                    screen = cv2.cvtColor(np.array(screenshot_pil), cv2.COLOR_RGB2BGR)
                    offset_x, offset_y = region_x, region_y
                else:
                    # 使用 Windows API 截取完整屏幕（支持多显示器）
                    try:
                        import win32gui
                        import win32ui
                        import win32con
                        
                        hdesktop = win32gui.GetDesktopWindow()
                        desktop_dc = win32gui.GetWindowDC(hdesktop)
                        img_dc = win32ui.CreateDCFromHandle(desktop_dc)
                        mem_dc = img_dc.CreateCompatibleDC()
                        
                        screenshot = win32ui.CreateBitmap()
                        screenshot.CreateCompatibleBitmap(img_dc, virtual_width, virtual_height)
                        mem_dc.SelectObject(screenshot)
                        mem_dc.BitBlt((0, 0), (virtual_width, virtual_height), img_dc, 
                                      (virtual_left, virtual_top), win32con.SRCCOPY)
                        
                        bmpinfo = screenshot.GetInfo()
                        bmpstr = screenshot.GetBitmapBits(True)
                        screen = np.frombuffer(bmpstr, dtype=np.uint8).reshape(
                            (bmpinfo['bmHeight'], bmpinfo['bmWidth'], 4))
                        screen = cv2.cvtColor(screen, cv2.COLOR_BGRA2BGR)
                        
                        mem_dc.DeleteDC()
                        win32gui.DeleteObject(screenshot.GetHandle())
                        win32gui.ReleaseDC(hdesktop, desktop_dc)
                        offset_x, offset_y = virtual_left, virtual_top
                        
                    except ImportError:
                        from PIL import ImageGrab
                        screenshot_pil = ImageGrab.grab(all_screens=True)
                        screen = cv2.cvtColor(np.array(screenshot_pil), cv2.COLOR_RGB2BGR)
                        offset_x, offset_y = 0, 0

                screen_gray = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)

                # 模板匹配
                result = cv2.matchTemplate(screen_gray, template_gray, cv2.TM_CCOEFF_NORMED)
                min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

                if max_val >= confidence:
                    # 找到匹配，根据点击位置计算实际坐标
                    img_left = offset_x + max_loc[0]
                    img_top = offset_y + max_loc[1]
                    img_right = img_left + w
                    img_bottom = img_top + h
                    
                    # 根据点击位置计算坐标
                    click_x, click_y = self._calculate_click_position(
                        click_position, img_left, img_top, img_right, img_bottom, w, h
                    )
                    
                    best_confidence = max_val
                    found = True
                    break

                await asyncio.sleep(0.3)

            if not found:
                return ModuleResult(
                    success=False, 
                    error=f"在 {wait_timeout} 秒内未找到匹配的图像（最高匹配度: {best_confidence:.2%}）"
                )

            # 执行点击 - 使用 SendInput API
            from ctypes import wintypes
            
            INPUT_MOUSE = 0
            MOUSEEVENTF_MOVE = 0x0001
            MOUSEEVENTF_ABSOLUTE = 0x8000
            MOUSEEVENTF_VIRTUALDESK = 0x4000  # 支持多显示器
            MOUSEEVENTF_LEFTDOWN = 0x0002
            MOUSEEVENTF_LEFTUP = 0x0004
            MOUSEEVENTF_RIGHTDOWN = 0x0008
            MOUSEEVENTF_RIGHTUP = 0x0010
            MOUSEEVENTF_MIDDLEDOWN = 0x0020
            MOUSEEVENTF_MIDDLEUP = 0x0040
            
            class MOUSEINPUT(ctypes.Structure):
                _fields_ = [
                    ("dx", wintypes.LONG),
                    ("dy", wintypes.LONG),
                    ("mouseData", wintypes.DWORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class INPUT(ctypes.Structure):
                _fields_ = [
                    ("type", wintypes.DWORD),
                    ("mi", MOUSEINPUT)
                ]

            # 使用虚拟桌面坐标系统
            abs_x = int((click_x - virtual_left) * 65535 / virtual_width)
            abs_y = int((click_y - virtual_top) * 65535 / virtual_height)

            # 选择按键事件
            if button == "left":
                down_event = MOUSEEVENTF_LEFTDOWN
                up_event = MOUSEEVENTF_LEFTUP
            elif button == "right":
                down_event = MOUSEEVENTF_RIGHTDOWN
                up_event = MOUSEEVENTF_RIGHTUP
            else:
                down_event = MOUSEEVENTF_MIDDLEDOWN
                up_event = MOUSEEVENTF_MIDDLEUP
            
            # 设置 DPI 感知
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except:
                    pass
            
            user32 = ctypes.windll.user32
            
            # 定义SendInput函数的参数类型（确保类型正确）
            user32.SendInput.argtypes = [
                wintypes.UINT,  # nInputs
                ctypes.POINTER(INPUT),  # pInputs
                ctypes.c_int  # cbSize
            ]
            user32.SendInput.restype = wintypes.UINT
            
            # 发送鼠标按键事件
            def send_mouse_event(event_flag):
                inp = INPUT()
                inp.type = INPUT_MOUSE
                inp.mi.dx = 0
                inp.mi.dy = 0
                inp.mi.mouseData = 0
                inp.mi.dwFlags = event_flag
                inp.mi.time = 0
                inp.mi.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                # 使用ctypes.pointer传递结构体指针
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    # 如果SendInput失败，记录错误但不中断执行
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result

            # 执行点击
            click_count = 2 if click_type == "double" else 1
            
            # 使用 SetCursorPos 移动鼠标（更精确）
            user32.SetCursorPos(int(click_x), int(click_y))
            await asyncio.sleep(0.02)
            
            for _ in range(click_count):
                send_mouse_event(down_event)
                await asyncio.sleep(0.05)
                send_mouse_event(up_event)
                if click_type == "double":
                    await asyncio.sleep(0.1)

            position_name = self._get_position_name(click_position)
            return ModuleResult(
                success=True, 
                message=f"已在图像{position_name} ({click_x}, {click_y}) 执行点击，匹配度: {best_confidence:.2%}",
                data={"x": click_x, "y": click_y, "confidence": best_confidence, "position": click_position}
            )

        except ImportError as e:
            missing = str(e).split("'")[1] if "'" in str(e) else "opencv-python/Pillow"
            return ModuleResult(
                success=False, 
                error=f"需要安装依赖库: pip install {missing}"
            )
        except Exception as e:
            return ModuleResult(success=False, error=f"点击图像失败: {str(e)}")
    
    def _calculate_click_position(self, position: str, left: int, top: int, right: int, bottom: int, w: int, h: int) -> tuple:
        """根据点击位置计算实际坐标"""
        if position == "center":
            return (left + w // 2, top + h // 2)
        elif position == "top-left":
            return (left + 5, top + 5)
        elif position == "top-right":
            return (right - 5, top + 5)
        elif position == "bottom-left":
            return (left + 5, bottom - 5)
        elif position == "bottom-right":
            return (right - 5, bottom - 5)
        elif position == "top":
            return (left + w // 2, top + 5)
        elif position == "bottom":
            return (left + w // 2, bottom - 5)
        elif position == "left":
            return (left + 5, top + h // 2)
        elif position == "right":
            return (right - 5, top + h // 2)
        elif position == "random":
            # 在图像范围内随机选择一个点（留5像素边距）
            margin = 5
            rand_x = random.randint(left + margin, right - margin)
            rand_y = random.randint(top + margin, bottom - margin)
            return (rand_x, rand_y)
        else:
            # 默认中心
            return (left + w // 2, top + h // 2)
    
    def _get_position_name(self, position: str) -> str:
        """获取位置的中文名称"""
        names = {
            "center": "中心",
            "top-left": "左上角",
            "top-right": "右上角",
            "bottom-left": "左下角",
            "bottom-right": "右下角",
            "top": "顶部",
            "bottom": "底部",
            "left": "左侧",
            "right": "右侧",
            "random": "随机位置"
        }
        return names.get(position, "中心")


@register_executor
class GetMousePositionExecutor(ModuleExecutor):
    """获取鼠标位置模块执行器"""

    @property
    def module_type(self) -> str:
        return "get_mouse_position"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes

        variable_name_x = config.get("variableNameX", "")
        variable_name_y = config.get("variableNameY", "")

        if not variable_name_x and not variable_name_y:
            return ModuleResult(success=False, error="至少需要指定一个变量名")

        try:
            class POINT(ctypes.Structure):
                _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]

            pt = POINT()
            ctypes.windll.user32.GetCursorPos(ctypes.byref(pt))

            if variable_name_x:
                context.set_variable(variable_name_x, pt.x)
            if variable_name_y:
                context.set_variable(variable_name_y, pt.y)

            return ModuleResult(
                success=True, 
                message=f"鼠标位置: ({pt.x}, {pt.y})",
                data={"x": pt.x, "y": pt.y}
            )

        except Exception as e:
            return ModuleResult(success=False, error=f"获取鼠标位置失败: {str(e)}")


@register_executor
class ScreenshotScreenExecutor(ModuleExecutor):
    """屏幕截图模块执行器"""

    @property
    def module_type(self) -> str:
        return "screenshot_screen"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        save_path = context.resolve_value(config.get("savePath", ""))
        file_name = context.resolve_value(config.get("fileName", ""))
        region = context.resolve_value(config.get("region", "full"))  # 支持变量引用
        variable_name = config.get("variableName", "")

        try:
            from PIL import ImageGrab
            import datetime

            # 生成文件名
            if not file_name:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"screenshot_{timestamp}.png"
            elif not file_name.lower().endswith('.png'):
                file_name += '.png'

            # 确定保存路径
            if save_path:
                Path(save_path).mkdir(parents=True, exist_ok=True)
                full_path = Path(save_path) / file_name
            else:
                screenshots_dir = Path("screenshots")
                screenshots_dir.mkdir(exist_ok=True)
                full_path = screenshots_dir / file_name

            # 截图
            if region == "custom":
                # 支持两点定位（x1,y1 到 x2,y2）
                x1 = to_int(config.get("x1", 0), 0, context)
                y1 = to_int(config.get("y1", 0), 0, context)
                x2 = to_int(config.get("x2", 800), 800, context)
                y2 = to_int(config.get("y2", 600), 600, context)
                
                # 确保坐标顺序正确（左上角到右下角）
                left = min(x1, x2)
                top = min(y1, y2)
                right = max(x1, x2)
                bottom = max(y1, y2)
                
                bbox = (left, top, right, bottom)
                screenshot = ImageGrab.grab(bbox=bbox)
                
                width = right - left
                height = bottom - top
                message = f"屏幕截图已保存: {full_path} (区域: {width}×{height})"
            else:
                screenshot = ImageGrab.grab()
                message = f"屏幕截图已保存: {full_path}"

            # 保存
            screenshot.save(str(full_path), "PNG")

            if variable_name:
                context.set_variable(variable_name, str(full_path))

            return ModuleResult(
                success=True,
                message=message,
                data=str(full_path)
            )

        except ImportError:
            return ModuleResult(success=False, error="需要安装 Pillow 库: pip install Pillow")
        except Exception as e:
            return ModuleResult(success=False, error=f"屏幕截图失败: {str(e)}")


@register_executor
class RenameFileExecutor(ModuleExecutor):
    """文件重命名模块执行器"""

    @property
    def module_type(self) -> str:
        return "rename_file"


    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        source_path = context.resolve_value(config.get("sourcePath", ""))
        new_name = context.resolve_value(config.get("newName", ""))
        variable_name = config.get("variableName", "")

        if not source_path:
            return ModuleResult(success=False, error="源文件路径不能为空")

        if not new_name:
            return ModuleResult(success=False, error="新文件名不能为空")

        try:
            source = Path(source_path)
            
            if not source.exists():
                return ModuleResult(success=False, error=f"源文件不存在: {source_path}")

            # 构建新路径（保持在同一目录）
            new_path = source.parent / new_name

            # 执行重命名
            source.rename(new_path)

            if variable_name:
                context.set_variable(variable_name, str(new_path))

            return ModuleResult(
                success=True,
                message=f"文件已重命名: {source.name} → {new_name}",
                data=str(new_path)
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限重命名该文件")
        except Exception as e:
            return ModuleResult(success=False, error=f"文件重命名失败: {str(e)}")


@register_executor
class NetworkCaptureExecutor(ModuleExecutor):
    """网络抓包模块执行器 - 支持浏览器抓包、系统抓包和代理抓包"""

    @property
    def module_type(self) -> str:
        return "network_capture"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        capture_mode = context.resolve_value(config.get("captureMode", "browser"))  # browser, system, proxy
        filter_type = context.resolve_value(config.get("filterType", "all"))  # all, img, media, m3u8
        search_keyword = context.resolve_value(config.get("searchKeyword", ""))
        capture_duration = to_float(config.get("captureDuration", 5), 5, context)  # 抓包时长(秒)
        variable_name = config.get("variableName", "")
        
        # 全局抓包特有参数
        target_process = context.resolve_value(config.get("targetProcess", ""))  # 目标进程名
        target_ports = context.resolve_value(config.get("targetPorts", ""))  # 目标端口，逗号分隔
        
        # 代理抓包特有参数
        proxy_port = to_int(config.get("proxyPort", 8888), 8888, context)

        if not variable_name:
            return ModuleResult(success=False, error="请指定存储变量名")

        try:
            if capture_mode == "proxy":
                # 代理抓包模式（用于模拟器/手机APP）
                return await self._capture_proxy(
                    context, variable_name, capture_duration,
                    filter_type, search_keyword, proxy_port
                )
            elif capture_mode == "system":
                # 全局系统抓包模式
                return await self._capture_system(
                    context, variable_name, capture_duration, 
                    search_keyword, target_process, target_ports
                )
            else:
                # 浏览器抓包模式
                return await self._capture_browser(
                    context, variable_name, capture_duration,
                    filter_type, search_keyword
                )

        except Exception as e:
            return ModuleResult(success=False, error=f"网络抓包启动失败: {str(e)}")
    
    async def _capture_proxy(self, context: 'ExecutionContext', variable_name: str,
                              capture_duration: int, filter_type: str, 
                              search_keyword: str, proxy_port: int) -> ModuleResult:
        """代理抓包模式 - 用于抓取模拟器/手机APP的HTTP请求"""
        from ..services.proxy_capture import proxy_capture_service
        import socket
        
        # 获取本机IP
        def get_local_ip():
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                s.connect(("8.8.8.8", 80))
                ip = s.getsockname()[0]
                s.close()
                return ip
            except:
                return "127.0.0.1"
        
        local_ip = get_local_ip()
        
        # 启动代理服务（如果未启动）
        if not proxy_capture_service.is_running:
            if not proxy_capture_service.start(proxy_port):
                return ModuleResult(
                    success=False, 
                    error="代理服务启动失败，请确保已安装 mitmproxy: pip install mitmproxy"
                )
        
        # 记录开始时间
        start_time = asyncio.get_event_loop().time()
        capture_start = asyncio.get_event_loop().time()
        
        # 清空之前的捕获（可选，这里选择不清空以便累积）
        # proxy_capture_service.clear_captured()
        
        # 等待抓包时间
        await asyncio.sleep(capture_duration)
        
        # 获取捕获的URL
        captured = proxy_capture_service.get_captured_urls(
            keyword=search_keyword,
            filter_type=filter_type,
            since=capture_start
        )
        
        # 提取URL列表
        urls = [item["url"] for item in captured]
        unique_urls = list(dict.fromkeys(urls))  # 去重保持顺序
        
        context.set_variable(variable_name, unique_urls)
        
        filter_desc = filter_type if filter_type != "all" else "全部"
        if search_keyword:
            filter_desc += f"，关键词={search_keyword}"
        
        return ModuleResult(
            success=True,
            message=f"代理抓包完成，捕获到 {len(unique_urls)} 个URL（过滤: {filter_desc}），已存入变量 {{{variable_name}}}。代理地址: {local_ip}:{proxy_port}",
            data={
                "status": "completed", 
                "mode": "proxy",
                "count": len(unique_urls), 
                "urls": unique_urls[:20],
                "proxy_ip": local_ip,
                "proxy_port": proxy_port
            }
        )
    
    async def _capture_browser(self, context: 'ExecutionContext', variable_name: str, 
                                capture_duration: int, filter_type: str, search_keyword: str) -> ModuleResult:
        """浏览器抓包模式 - 使用 Playwright 监听页面请求"""
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面，浏览器抓包需要先打开网页")

        await context.switch_to_latest_page()
        
        # 存储捕获到的请求URL
        captured_urls = []
        
        # 定义资源类型过滤
        def should_capture(request) -> bool:
            url = request.url
            resource_type = request.resource_type
            
            # 根据过滤类型筛选
            if filter_type == "img":
                if resource_type not in ["image"]:
                    return False
            elif filter_type == "media":
                if resource_type not in ["media", "video", "audio"]:
                    return False
            # all 类型不过滤
            
            # 模糊搜索筛选
            if search_keyword:
                if search_keyword.lower() not in url.lower():
                    return False
            
            return True
        
        # 请求处理函数
        def on_request(request):
            if should_capture(request):
                captured_urls.append(request.url)
        
        # 注册请求监听器
        context.page.on("request", on_request)
        
        # 等待抓包时间（阻塞）
        await asyncio.sleep(capture_duration)
        
        # 移除监听器
        context.page.remove_listener("request", on_request)
        
        # 去重并存储结果
        unique_urls = list(dict.fromkeys(captured_urls))
        context.set_variable(variable_name, unique_urls)
        
        return ModuleResult(
            success=True,
            message=f"浏览器抓包完成，捕获到 {len(unique_urls)} 个URL，已存入变量 {{{variable_name}}}",
            data={"status": "completed", "mode": "browser", "count": len(unique_urls), "urls": unique_urls[:10]}
        )
    
    async def _capture_system(self, context: 'ExecutionContext', variable_name: str,
                               capture_duration: int, search_keyword: str,
                               target_process: str, target_ports: str) -> ModuleResult:
        """全局系统抓包模式 - 监控系统网络连接"""
        import psutil
        
        # 解析目标端口
        port_filter = set()
        if target_ports:
            for p in target_ports.replace('，', ',').split(','):
                p = p.strip()
                if p.isdigit():
                    port_filter.add(int(p))
        
        # 存储捕获到的连接信息
        captured_connections = []
        seen_connections = set()  # 用于去重
        
        def get_process_name(pid: int) -> str:
            """获取进程名"""
            try:
                proc = psutil.Process(pid)
                return proc.name()
            except:
                return ""
        
        def capture_connections():
            """捕获当前网络连接"""
            try:
                connections = psutil.net_connections(kind='inet')
                for conn in connections:
                    # 只关注已建立的连接
                    if conn.status != 'ESTABLISHED':
                        continue
                    
                    # 获取远程地址
                    if not conn.raddr:
                        continue
                    
                    remote_ip = conn.raddr.ip
                    remote_port = conn.raddr.port
                    local_port = conn.laddr.port if conn.laddr else 0
                    pid = conn.pid or 0
                    
                    # 端口过滤
                    if port_filter:
                        if remote_port not in port_filter and local_port not in port_filter:
                            continue
                    
                    # 进程过滤
                    proc_name = get_process_name(pid) if pid else ""
                    if target_process:
                        if target_process.lower() not in proc_name.lower():
                            continue
                    
                    # 构建连接标识用于去重
                    conn_key = f"{remote_ip}:{remote_port}:{pid}"
                    if conn_key in seen_connections:
                        continue
                    seen_connections.add(conn_key)
                    
                    # 构建连接信息
                    conn_info = {
                        "remote_ip": remote_ip,
                        "remote_port": remote_port,
                        "local_port": local_port,
                        "pid": pid,
                        "process": proc_name,
                        "address": f"{remote_ip}:{remote_port}"
                    }
                    
                    # 关键词过滤
                    if search_keyword:
                        search_str = f"{remote_ip}:{remote_port} {proc_name}".lower()
                        if search_keyword.lower() not in search_str:
                            continue
                    
                    captured_connections.append(conn_info)
            except Exception as e:
                print(f"[DEBUG] 捕获连接异常: {e}")
        
        # 创建异步任务来持续抓包
        async def capture_task():
            import time
            start_time = time.time()
            duration_sec = capture_duration
            
            while time.time() - start_time < duration_sec:
                # 在线程池中执行连接捕获
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(None, capture_connections)
                await asyncio.sleep(0.5)  # 每500ms扫描一次
            
            # 存储结果
            context.set_variable(variable_name, captured_connections)
        
        # 等待抓包完成（阻塞）
        await capture_task()
        
        filter_desc = []
        if target_process:
            filter_desc.append(f"进程={target_process}")
        if target_ports:
            filter_desc.append(f"端口={target_ports}")
        if search_keyword:
            filter_desc.append(f"关键词={search_keyword}")
        filter_str = "，".join(filter_desc) if filter_desc else "无"
        
        return ModuleResult(
            success=True,
            message=f"全局抓包完成，过滤条件: {filter_str}，捕获到 {len(captured_connections)} 个连接，已存入变量 {{{variable_name}}}",
            data={"status": "completed", "mode": "system", "count": len(captured_connections), "connections": captured_connections[:10]}
        )



@register_executor
class ListFilesExecutor(ModuleExecutor):
    """获取文件列表模块执行器"""

    @property
    def module_type(self) -> str:
        return "list_files"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import fnmatch
        
        folder_path = context.resolve_value(config.get("folderPath", ""))
        list_type = context.resolve_value(config.get("listType", "files"))  # 支持变量引用: files, folders, all
        include_extension_raw = config.get("includeExtension", True)
        # 支持变量引用
        if isinstance(include_extension_raw, str):
            include_extension_raw = context.resolve_value(include_extension_raw)
        include_extension = include_extension_raw in [True, 'true', 'True', '1', 1]
        filter_pattern = context.resolve_value(config.get("filterPattern", ""))
        variable_name = config.get("resultVariable", "file_list")

        if not folder_path:
            return ModuleResult(success=False, error="文件夹路径不能为空")

        try:
            folder = Path(folder_path)
            
            if not folder.exists():
                return ModuleResult(success=False, error=f"文件夹不存在: {folder_path}")
            
            if not folder.is_dir():
                return ModuleResult(success=False, error=f"路径不是文件夹: {folder_path}")

            result_list = []
            
            # 解析过滤模式
            patterns = []
            if filter_pattern:
                patterns = [p.strip() for p in filter_pattern.split(';') if p.strip()]
            
            for item in folder.iterdir():
                # 根据类型过滤
                if list_type == "files" and not item.is_file():
                    continue
                if list_type == "folders" and not item.is_dir():
                    continue
                
                name = item.name
                
                # 应用过滤模式
                if patterns and item.is_file():
                    matched = any(fnmatch.fnmatch(name.lower(), p.lower()) for p in patterns)
                    if not matched:
                        continue
                
                # 处理文件名格式
                if item.is_file() and not include_extension:
                    name = item.stem
                
                result_list.append(name)
            
            # 排序
            result_list.sort()
            
            if variable_name:
                context.set_variable(variable_name, result_list)

            return ModuleResult(
                success=True,
                message=f"获取到 {len(result_list)} 个项目",
                data=result_list
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限访问该文件夹")
        except Exception as e:
            return ModuleResult(success=False, error=f"获取文件列表失败: {str(e)}")


@register_executor
class CopyFileExecutor(ModuleExecutor):
    """复制文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "copy_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import shutil
        
        source_path = context.resolve_value(config.get("sourcePath", ""))
        target_path = context.resolve_value(config.get("targetPath", ""))
        overwrite_raw = config.get("overwrite", True)
        # 支持变量引用
        if isinstance(overwrite_raw, str):
            overwrite_raw = context.resolve_value(overwrite_raw)
        overwrite = overwrite_raw in [True, 'true', 'True', '1', 1]
        variable_name = config.get("resultVariable", "copied_path")

        if not source_path:
            return ModuleResult(success=False, error="源文件路径不能为空")
        
        if not target_path:
            return ModuleResult(success=False, error="目标路径不能为空")

        try:
            source = Path(source_path)
            target = Path(target_path)
            
            if not source.exists():
                return ModuleResult(success=False, error=f"源文件不存在: {source_path}")
            
            # 如果目标是文件夹，则保持原文件名
            if target.is_dir() or str(target_path).endswith(('/', '\\')):
                target.mkdir(parents=True, exist_ok=True)
                target = target / source.name
            else:
                target.parent.mkdir(parents=True, exist_ok=True)
            
            # 检查是否覆盖
            if target.exists() and not overwrite:
                return ModuleResult(
                    success=True,
                    message=f"目标文件已存在，已跳过: {target}",
                    data=str(target)
                )
            
            # 执行复制
            shutil.copy2(source, target)
            
            if variable_name:
                context.set_variable(variable_name, str(target))

            return ModuleResult(
                success=True,
                message=f"文件已复制: {source.name} → {target}",
                data=str(target)
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限执行复制操作")
        except Exception as e:
            return ModuleResult(success=False, error=f"复制文件失败: {str(e)}")


@register_executor
class MoveFileExecutor(ModuleExecutor):
    """移动文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "move_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import shutil
        
        source_path = context.resolve_value(config.get("sourcePath", ""))
        target_path = context.resolve_value(config.get("targetPath", ""))
        overwrite_raw = config.get("overwrite", True)
        # 支持变量引用
        if isinstance(overwrite_raw, str):
            overwrite_raw = context.resolve_value(overwrite_raw)
        overwrite = overwrite_raw in [True, 'true', 'True', '1', 1]
        variable_name = config.get("resultVariable", "moved_path")

        if not source_path:
            return ModuleResult(success=False, error="源文件路径不能为空")
        
        if not target_path:
            return ModuleResult(success=False, error="目标路径不能为空")

        try:
            source = Path(source_path)
            target = Path(target_path)
            
            if not source.exists():
                return ModuleResult(success=False, error=f"源文件不存在: {source_path}")
            
            # 如果目标是文件夹，则保持原文件名
            if target.is_dir() or str(target_path).endswith(('/', '\\')):
                target.mkdir(parents=True, exist_ok=True)
                target = target / source.name
            else:
                target.parent.mkdir(parents=True, exist_ok=True)
            
            # 检查是否覆盖
            if target.exists():
                if not overwrite:
                    return ModuleResult(
                        success=True,
                        message=f"目标文件已存在，已跳过: {target}",
                        data=str(target)
                    )
                target.unlink()
            
            # 执行移动
            shutil.move(str(source), str(target))
            
            if variable_name:
                context.set_variable(variable_name, str(target))

            return ModuleResult(
                success=True,
                message=f"文件已移动: {source.name} → {target}",
                data=str(target)
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限执行移动操作")
        except Exception as e:
            return ModuleResult(success=False, error=f"移动文件失败: {str(e)}")


@register_executor
class DeleteFileExecutor(ModuleExecutor):
    """删除文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "delete_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import shutil
        
        file_path = context.resolve_value(config.get("filePath", ""))
        delete_type = context.resolve_value(config.get("deleteType", "file"))  # 支持变量引用: file, folder

        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")

        try:
            path = Path(file_path)
            
            if not path.exists():
                return ModuleResult(
                    success=True,
                    message=f"文件/文件夹不存在，无需删除: {file_path}"
                )
            
            if delete_type == "folder":
                if path.is_dir():
                    shutil.rmtree(path)
                    return ModuleResult(
                        success=True,
                        message=f"文件夹已删除: {file_path}"
                    )
                else:
                    return ModuleResult(success=False, error=f"路径不是文件夹: {file_path}")
            else:
                if path.is_file():
                    path.unlink()
                    return ModuleResult(
                        success=True,
                        message=f"文件已删除: {file_path}"
                    )
                else:
                    return ModuleResult(success=False, error=f"路径不是文件: {file_path}")

        except PermissionError:
            return ModuleResult(success=False, error="没有权限删除该文件/文件夹")
        except Exception as e:
            return ModuleResult(success=False, error=f"删除失败: {str(e)}")


@register_executor
class CreateFolderExecutor(ModuleExecutor):
    """创建文件夹模块执行器"""

    @property
    def module_type(self) -> str:
        return "create_folder"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        folder_path = context.resolve_value(config.get("folderPath", ""))
        variable_name = config.get("resultVariable", "folder_path")

        if not folder_path:
            return ModuleResult(success=False, error="文件夹路径不能为空")

        try:
            folder = Path(folder_path)
            
            if folder.exists():
                if folder.is_dir():
                    if variable_name:
                        context.set_variable(variable_name, str(folder))
                    return ModuleResult(
                        success=True,
                        message=f"文件夹已存在: {folder_path}",
                        data=str(folder)
                    )
                else:
                    return ModuleResult(success=False, error=f"路径已存在且不是文件夹: {folder_path}")
            
            # 创建文件夹（包括父目录）
            folder.mkdir(parents=True, exist_ok=True)
            
            if variable_name:
                context.set_variable(variable_name, str(folder))

            return ModuleResult(
                success=True,
                message=f"文件夹已创建: {folder_path}",
                data=str(folder)
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限创建文件夹")
        except Exception as e:
            return ModuleResult(success=False, error=f"创建文件夹失败: {str(e)}")


@register_executor
class FileExistsExecutor(ModuleExecutor):
    """文件是否存在模块执行器"""

    @property
    def module_type(self) -> str:
        return "file_exists"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        file_path = context.resolve_value(config.get("filePath", ""))
        variable_name = config.get("resultVariable", "exists")

        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")

        try:
            path = Path(file_path)
            exists = path.exists()
            
            if variable_name:
                context.set_variable(variable_name, exists)

            return ModuleResult(
                success=True,
                message=f"{'存在' if exists else '不存在'}: {file_path}",
                data=exists
            )

        except Exception as e:
            return ModuleResult(success=False, error=f"检查文件存在失败: {str(e)}")


@register_executor
class GetFileInfoExecutor(ModuleExecutor):
    """获取文件信息模块执行器"""

    @property
    def module_type(self) -> str:
        return "get_file_info"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import os
        from datetime import datetime
        
        file_path = context.resolve_value(config.get("filePath", ""))
        variable_name = config.get("resultVariable", "file_info")

        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")

        try:
            path = Path(file_path)
            
            if not path.exists():
                return ModuleResult(success=False, error=f"文件不存在: {file_path}")
            
            stat = path.stat()
            
            file_info = {
                "name": path.name,
                "path": str(path.absolute()),
                "size": stat.st_size,
                "extension": path.suffix.lower() if path.is_file() else "",
                "created_time": datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S"),
                "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "is_file": path.is_file(),
                "is_folder": path.is_dir(),
            }
            
            if variable_name:
                context.set_variable(variable_name, file_info)

            size_str = self._format_size(stat.st_size)
            return ModuleResult(
                success=True,
                message=f"{path.name} ({size_str})",
                data=file_info
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限访问该文件")
        except Exception as e:
            return ModuleResult(success=False, error=f"获取文件信息失败: {str(e)}")
    
    def _format_size(self, size: int) -> str:
        """格式化文件大小"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} TB"


@register_executor
class ReadTextFileExecutor(ModuleExecutor):
    """读取文本文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "read_text_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        file_path = context.resolve_value(config.get("filePath", ""))
        encoding = context.resolve_value(config.get("encoding", "utf-8"))  # 支持变量引用
        variable_name = config.get("resultVariable", "file_content")

        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")

        try:
            path = Path(file_path)
            
            if not path.exists():
                return ModuleResult(success=False, error=f"文件不存在: {file_path}")
            
            if not path.is_file():
                return ModuleResult(success=False, error=f"路径不是文件: {file_path}")
            
            # 读取文件内容
            with open(path, 'r', encoding=encoding) as f:
                content = f.read()
            
            if variable_name:
                context.set_variable(variable_name, content)

            # 显示内容预览
            preview = content[:100] + "..." if len(content) > 100 else content
            preview = preview.replace('\n', '\\n')
            
            return ModuleResult(
                success=True,
                message=f"已读取 {len(content)} 字符: {preview}",
                data=content
            )

        except UnicodeDecodeError:
            return ModuleResult(success=False, error=f"文件编码错误，请尝试其他编码格式（当前: {encoding}）")
        except PermissionError:
            return ModuleResult(success=False, error="没有权限读取该文件")
        except Exception as e:
            return ModuleResult(success=False, error=f"读取文件失败: {str(e)}")


@register_executor
class WriteTextFileExecutor(ModuleExecutor):
    """写入文本文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "write_text_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        file_path = context.resolve_value(config.get("filePath", ""))
        content = context.resolve_value(config.get("content", ""))
        encoding = context.resolve_value(config.get("encoding", "utf-8"))  # 支持变量引用
        write_mode = context.resolve_value(config.get("writeMode", "overwrite"))  # 支持变量引用: overwrite, append
        variable_name = config.get("resultVariable", "write_path")

        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")

        try:
            path = Path(file_path)
            
            # 确保父目录存在
            path.parent.mkdir(parents=True, exist_ok=True)
            
            # 确定写入模式
            mode = 'a' if write_mode == 'append' else 'w'
            
            # 写入文件
            with open(path, mode, encoding=encoding) as f:
                f.write(content)
            
            if variable_name:
                context.set_variable(variable_name, str(path))

            action = "追加" if write_mode == 'append' else "写入"
            return ModuleResult(
                success=True,
                message=f"已{action} {len(content)} 字符到: {path.name}",
                data=str(path)
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限写入该文件")
        except Exception as e:
            return ModuleResult(success=False, error=f"写入文件失败: {str(e)}")



@register_executor
class RenameFolderExecutor(ModuleExecutor):
    """文件夹重命名模块执行器"""

    @property
    def module_type(self) -> str:
        return "rename_folder"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        source_path = context.resolve_value(config.get("sourcePath", ""))
        new_name = context.resolve_value(config.get("newName", ""))
        variable_name = config.get("resultVariable", "new_folder_path")

        if not source_path:
            return ModuleResult(success=False, error="源文件夹路径不能为空")

        if not new_name:
            return ModuleResult(success=False, error="新文件夹名不能为空")

        try:
            source = Path(source_path)
            
            if not source.exists():
                return ModuleResult(success=False, error=f"源文件夹不存在: {source_path}")
            
            if not source.is_dir():
                return ModuleResult(success=False, error=f"路径不是文件夹: {source_path}")

            # 构建新路径（保持在同一父目录）
            new_path = source.parent / new_name

            # 检查目标是否已存在
            if new_path.exists():
                return ModuleResult(success=False, error=f"目标文件夹已存在: {new_path}")

            # 执行重命名
            source.rename(new_path)

            if variable_name:
                context.set_variable(variable_name, str(new_path))

            return ModuleResult(
                success=True,
                message=f"文件夹已重命名: {source.name} → {new_name}",
                data=str(new_path)
            )

        except PermissionError:
            return ModuleResult(success=False, error="没有权限重命名该文件夹")
        except Exception as e:
            return ModuleResult(success=False, error=f"文件夹重命名失败: {str(e)}")


@register_executor
class MacroRecorderExecutor(ModuleExecutor):
    """宏录制器模块执行器 - 录制并回放鼠标和键盘操作"""

    @property
    def module_type(self) -> str:
        return "macro_recorder"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes
        import json

        # 获取配置
        recorded_data = config.get("recordedData", "")  # JSON格式的录制数据
        play_speed = to_float(config.get("playSpeed", 1.0), 1.0, context)  # 播放速度倍率
        repeat_count = to_int(config.get("repeatCount", 1), 1, context)  # 重复次数
        
        # 播放选项
        play_mouse_move = config.get("playMouseMove", True)  # 播放鼠标移动轨迹
        play_mouse_click = config.get("playMouseClick", True)  # 播放鼠标点击
        play_keyboard = config.get("playKeyboard", True)  # 播放键盘操作
        use_relative_position = config.get("useRelativePosition", False)  # 使用相对位置
        
        # 相对位置的基准点（如果启用相对位置）
        base_x = to_int(config.get("baseX", 0), 0, context)
        base_y = to_int(config.get("baseY", 0), 0, context)

        if not recorded_data:
            return ModuleResult(success=False, error="没有录制数据，请先录制操作")

        try:
            # 解析录制数据
            if isinstance(recorded_data, str):
                actions = json.loads(recorded_data)
            else:
                actions = recorded_data
            
            if not actions or not isinstance(actions, list):
                return ModuleResult(success=False, error="录制数据格式无效")

            from ctypes import wintypes
            
            # Windows API
            user32 = ctypes.windll.user32
            
            # 设置进程为 DPI 感知，确保坐标与录制时一致
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
            except:
                try:
                    user32.SetProcessDPIAware()
                except:
                    pass
            
            # SendInput 结构体定义
            INPUT_MOUSE = 0
            INPUT_KEYBOARD = 1
            MOUSEEVENTF_MOVE = 0x0001
            MOUSEEVENTF_LEFTDOWN = 0x0002
            MOUSEEVENTF_LEFTUP = 0x0004
            MOUSEEVENTF_RIGHTDOWN = 0x0008
            MOUSEEVENTF_RIGHTUP = 0x0010
            MOUSEEVENTF_MIDDLEDOWN = 0x0020
            MOUSEEVENTF_MIDDLEUP = 0x0040
            MOUSEEVENTF_WHEEL = 0x0800
            MOUSEEVENTF_VIRTUALDESK = 0x4000
            MOUSEEVENTF_ABSOLUTE = 0x8000
            KEYEVENTF_KEYUP = 0x0002
            KEYEVENTF_UNICODE = 0x0004
            
            class MOUSEINPUT(ctypes.Structure):
                _fields_ = [
                    ("dx", wintypes.LONG),
                    ("dy", wintypes.LONG),
                    ("mouseData", wintypes.DWORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class KEYBDINPUT(ctypes.Structure):
                _fields_ = [
                    ("wVk", wintypes.WORD),
                    ("wScan", wintypes.WORD),
                    ("dwFlags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            class INPUT_UNION(ctypes.Union):
                _fields_ = [("mi", MOUSEINPUT), ("ki", KEYBDINPUT)]
            
            class INPUT(ctypes.Structure):
                _fields_ = [
                    ("type", wintypes.DWORD),
                    ("union", INPUT_UNION)
                ]
            
            # 定义SendInput函数的参数类型（确保类型正确）
            user32.SendInput.argtypes = [
                wintypes.UINT,  # nInputs
                ctypes.POINTER(INPUT),  # pInputs
                ctypes.c_int  # cbSize
            ]
            user32.SendInput.restype = wintypes.UINT
            
            # 获取虚拟屏幕尺寸（用于坐标转换）
            SM_XVIRTUALSCREEN = 76
            SM_YVIRTUALSCREEN = 77
            SM_CXVIRTUALSCREEN = 78
            SM_CYVIRTUALSCREEN = 79
            
            virtual_left = user32.GetSystemMetrics(SM_XVIRTUALSCREEN)
            virtual_top = user32.GetSystemMetrics(SM_YVIRTUALSCREEN)
            virtual_width = user32.GetSystemMetrics(SM_CXVIRTUALSCREEN)
            virtual_height = user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)
            
            # 安装低级鼠标钩子来清除 LLMHF_INJECTED 标志
            # 这样可以让模拟的鼠标输入看起来像真实的硬件输入
            WH_MOUSE_LL = 14
            LLMHF_INJECTED = 0x00000001
            
            class MSLLHOOKSTRUCT(ctypes.Structure):
                _fields_ = [
                    ("pt", wintypes.POINT),
                    ("mouseData", wintypes.DWORD),
                    ("flags", wintypes.DWORD),
                    ("time", wintypes.DWORD),
                    ("dwExtraInfo", ctypes.POINTER(ctypes.c_ulong))
                ]
            
            # 钩子回调函数类型
            HOOKPROC = ctypes.WINFUNCTYPE(ctypes.c_long, ctypes.c_int, wintypes.WPARAM, wintypes.LPARAM)
            
            # 钩子句柄
            mouse_hook = None
            
            # 钩子回调函数 - 清除 LLMHF_INJECTED 标志
            @HOOKPROC
            def mouse_hook_proc(nCode, wParam, lParam):
                if nCode >= 0:
                    # 获取钩子结构体
                    hook_struct = ctypes.cast(lParam, ctypes.POINTER(MSLLHOOKSTRUCT)).contents
                    # 清除 LLMHF_INJECTED 标志
                    # 注意：这个修改可能不会传播到其他钩子，但值得一试
                    if hook_struct.flags & LLMHF_INJECTED:
                        hook_struct.flags &= ~LLMHF_INJECTED
                return user32.CallNextHookEx(mouse_hook, nCode, wParam, lParam)
            
            # 安装钩子
            try:
                mouse_hook = user32.SetWindowsHookExW(WH_MOUSE_LL, mouse_hook_proc, None, 0)
            except:
                mouse_hook = None
            
            # mouse_event 常量
            ME_MOVE = 0x0001
            ME_ABSOLUTE = 0x8000
            ME_LEFTDOWN = 0x0002
            ME_LEFTUP = 0x0004
            ME_RIGHTDOWN = 0x0008
            ME_RIGHTUP = 0x0010
            ME_MIDDLEDOWN = 0x0020
            ME_MIDDLEUP = 0x0040
            ME_WHEEL = 0x0800
            
            # 移动鼠标 - 使用 SetCursorPos（直接设置光标位置，更可靠）
            def move_mouse(x, y):
                x = int(x)
                y = int(y)
                # 使用 SetCursorPos 直接设置光标位置
                user32.SetCursorPos(x, y)
            
            # 发送鼠标按键事件 - 先移动到位置再点击
            def send_mouse_button(event_flag, x=None, y=None):
                # 如果提供了坐标，先移动到该位置
                if x is not None and y is not None:
                    user32.SetCursorPos(int(x), int(y))
                
                # 映射 SendInput 标志到 mouse_event 标志
                me_flag = 0
                if event_flag == MOUSEEVENTF_LEFTDOWN:
                    me_flag = ME_LEFTDOWN
                elif event_flag == MOUSEEVENTF_LEFTUP:
                    me_flag = ME_LEFTUP
                elif event_flag == MOUSEEVENTF_RIGHTDOWN:
                    me_flag = ME_RIGHTDOWN
                elif event_flag == MOUSEEVENTF_RIGHTUP:
                    me_flag = ME_RIGHTUP
                elif event_flag == MOUSEEVENTF_MIDDLEDOWN:
                    me_flag = ME_MIDDLEDOWN
                elif event_flag == MOUSEEVENTF_MIDDLEUP:
                    me_flag = ME_MIDDLEUP
                
                # 使用 mouse_event 发送按键事件
                user32.mouse_event(me_flag, 0, 0, 0, 0)
            
            def send_mouse_scroll(delta):
                # 使用 mouse_event 发送滚轮事件
                user32.mouse_event(ME_WHEEL, 0, 0, delta, 0)
            
            def send_key(vk_code, is_up=False):
                inp = INPUT()
                inp.type = INPUT_KEYBOARD
                inp.union.ki.wVk = vk_code
                inp.union.ki.wScan = user32.MapVirtualKeyW(vk_code, 0)
                inp.union.ki.dwFlags = KEYEVENTF_KEYUP if is_up else 0
                inp.union.ki.time = 0
                inp.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result
            
            def send_unicode_char(char):
                # 按下
                inp_down = INPUT()
                inp_down.type = INPUT_KEYBOARD
                inp_down.union.ki.wVk = 0
                inp_down.union.ki.wScan = ord(char)
                inp_down.union.ki.dwFlags = KEYEVENTF_UNICODE
                inp_down.union.ki.time = 0
                inp_down.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp_down), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                # 释放
                inp_up = INPUT()
                inp_up.type = INPUT_KEYBOARD
                inp_up.union.ki.wVk = 0
                inp_up.union.ki.wScan = ord(char)
                inp_up.union.ki.dwFlags = KEYEVENTF_UNICODE | KEYEVENTF_KEYUP
                inp_up.union.ki.time = 0
                inp_up.union.ki.dwExtraInfo = ctypes.pointer(ctypes.c_ulong(0))
                result = user32.SendInput(1, ctypes.pointer(inp_up), ctypes.sizeof(INPUT))
                if result == 0:
                    error_code = ctypes.get_last_error()
                    print(f"SendInput failed with error code: {error_code}")
                return result

            # 如果使用相对位置，获取当前鼠标位置作为基准
            if use_relative_position:
                class POINT(ctypes.Structure):
                    _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]
                pt = POINT()
                user32.GetCursorPos(ctypes.byref(pt))
                # 计算偏移量（当前位置 - 录制时的基准位置）
                offset_x = pt.x - base_x
                offset_y = pt.y - base_y
            else:
                offset_x = 0
                offset_y = 0

            total_actions = len(actions)

            for repeat in range(repeat_count):
                if repeat_count > 1:
                    await context.send_progress(f"🔄 第 {repeat + 1}/{repeat_count} 次播放...")

                last_time = 0
                for i, action in enumerate(actions):
                    action_type = action.get("type")
                    timestamp = action.get("time", 0)
                    
                    # 计算延迟时间（考虑播放速度）
                    if i > 0 and timestamp > last_time:
                        delay = (timestamp - last_time) / 1000 / play_speed
                        # 使用更精确的延迟，最小延迟0.001秒
                        if delay > 0.001:
                            await asyncio.sleep(delay)
                    last_time = timestamp

                    # 执行动作
                    if action_type == "mouse_move" and play_mouse_move:
                        x = action.get("x", 0) + offset_x
                        y = action.get("y", 0) + offset_y
                        try:
                            move_mouse(x, y)
                        except Exception as e:
                            print(f"鼠标移动失败: {e}")
                        # 使用同步延迟，确保事件连续发送
                        import time as time_module
                        time_module.sleep(0.002)  # 2ms 延迟

                    elif action_type == "mouse_click" and play_mouse_click:
                        x = action.get("x", 0) + offset_x
                        y = action.get("y", 0) + offset_y
                        button = action.get("button", "left")
                        pressed = action.get("pressed", True)
                        
                        # 先移动到位置
                        try:
                            move_mouse(x, y)
                        except Exception as e:
                            print(f"鼠标移动失败: {e}")
                        
                        # 短暂延迟，模拟真实操作
                        import time as time_module
                        time_module.sleep(0.01)  # 10ms
                        
                        # 发送按键事件
                        if button == "left":
                            event = MOUSEEVENTF_LEFTDOWN if pressed else MOUSEEVENTF_LEFTUP
                        elif button == "right":
                            event = MOUSEEVENTF_RIGHTDOWN if pressed else MOUSEEVENTF_RIGHTUP
                        elif button == "middle":
                            event = MOUSEEVENTF_MIDDLEDOWN if pressed else MOUSEEVENTF_MIDDLEUP
                        else:
                            continue
                        send_mouse_button(event, x, y)
                        
                        # 点击后短暂延迟
                        time_module.sleep(0.01)  # 10ms

                    elif action_type == "mouse_scroll" and play_mouse_click:
                        delta = action.get("delta", 0)
                        send_mouse_scroll(delta)

                    elif action_type == "key_press" and play_keyboard:
                        key_code = action.get("keyCode", 0)
                        pressed = action.get("pressed", True)
                        if key_code > 0:
                            send_key(key_code, is_up=not pressed)

                    elif action_type == "key_char" and play_keyboard:
                        char = action.get("char", "")
                        if char:
                            send_unicode_char(char)

            # 统计信息
            move_count = sum(1 for a in actions if a.get("type") == "mouse_move")
            click_count = sum(1 for a in actions if a.get("type") == "mouse_click")
            key_count = sum(1 for a in actions if a.get("type") in ("key_press", "key_char"))
            
            message = f"宏播放完成: {total_actions}个动作"
            if repeat_count > 1:
                message += f" × {repeat_count}次"
            details = []
            if move_count > 0 and play_mouse_move:
                details.append(f"移动{move_count}次")
            if click_count > 0 and play_mouse_click:
                details.append(f"点击{click_count}次")
            if key_count > 0 and play_keyboard:
                details.append(f"按键{key_count}次")
            if details:
                message += f" ({', '.join(details)})"

            # 卸载鼠标钩子
            if mouse_hook:
                user32.UnhookWindowsHookEx(mouse_hook)

            return ModuleResult(
                success=True,
                message=message,
                data={
                    "total_actions": total_actions,
                    "repeat_count": repeat_count,
                    "move_count": move_count,
                    "click_count": click_count,
                    "key_count": key_count
                }
            )

        except json.JSONDecodeError:
            # 卸载鼠标钩子
            if 'mouse_hook' in dir() and mouse_hook:
                user32.UnhookWindowsHookEx(mouse_hook)
            return ModuleResult(success=False, error="录制数据JSON格式无效")
        except Exception as e:
            # 卸载鼠标钩子
            if 'mouse_hook' in dir() and mouse_hook:
                user32.UnhookWindowsHookEx(mouse_hook)
            return ModuleResult(success=False, error=f"宏播放失败: {str(e)}")


@register_executor
class ExportLogExecutor(ModuleExecutor):
    """导出日志模块执行器 - 将工作流执行日志导出到文件"""
    
    @property
    def module_type(self) -> str:
        return "export_log"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        output_path = context.resolve_value(config.get('outputPath', ''))
        log_format = config.get('logFormat', 'txt')  # txt, json, csv
        include_timestamp = config.get('includeTimestamp', True)
        include_level = config.get('includeLevel', True)
        include_duration = config.get('includeDuration', True)
        result_variable = config.get('resultVariable', '')
        
        if not output_path:
            # 默认保存到项目根目录的logs文件夹
            import datetime
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = Path(__file__).parent.parent.parent.parent / 'logs'
            output_dir.mkdir(exist_ok=True)
            output_path = str(output_dir / f'workflow_log_{timestamp}.{log_format}')
        
        try:
            # 获取日志数据
            logs = context.get_logs() if hasattr(context, 'get_logs') else []
            
            # 如果context没有get_logs方法，尝试从其他地方获取
            if not logs and hasattr(context, '_logs'):
                logs = context._logs
            
            # 确保输出目录存在
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)
            
            if log_format == 'json':
                result = self._export_json(logs, output_path, include_timestamp, include_level, include_duration)
            elif log_format == 'csv':
                result = self._export_csv(logs, output_path, include_timestamp, include_level, include_duration)
            else:
                result = self._export_txt(logs, output_path, include_timestamp, include_level, include_duration)
            
            if result_variable:
                context.set_variable(result_variable, result)
            
            return ModuleResult(
                success=True,
                message=f"已导出 {result['log_count']} 条日志到 {output_path}",
                data=result
            )
        except Exception as e:
            return ModuleResult(success=False, error=f"导出日志失败: {str(e)}")
    
    def _export_txt(self, logs: list, output_path: str, include_timestamp: bool, 
                    include_level: bool, include_duration: bool) -> dict:
        """导出为TXT格式"""
        lines = []
        for log in logs:
            parts = []
            if include_timestamp and 'timestamp' in log:
                parts.append(f"[{log['timestamp']}]")
            if include_level and 'level' in log:
                parts.append(f"[{log['level'].upper()}]")
            parts.append(log.get('message', ''))
            if include_duration and 'duration' in log and log['duration']:
                parts.append(f"({log['duration']:.2f}ms)")
            lines.append(' '.join(parts))
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return {
            "output_path": output_path,
            "log_count": len(logs),
            "format": "txt",
            "file_size": Path(output_path).stat().st_size
        }
    
    def _export_json(self, logs: list, output_path: str, include_timestamp: bool,
                     include_level: bool, include_duration: bool) -> dict:
        """导出为JSON格式"""
        export_logs = []
        for log in logs:
            export_log = {"message": log.get('message', '')}
            if include_timestamp and 'timestamp' in log:
                export_log['timestamp'] = log['timestamp']
            if include_level and 'level' in log:
                export_log['level'] = log['level']
            if include_duration and 'duration' in log:
                export_log['duration'] = log['duration']
            if 'nodeId' in log:
                export_log['nodeId'] = log['nodeId']
            export_logs.append(export_log)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_logs, f, ensure_ascii=False, indent=2)
        
        return {
            "output_path": output_path,
            "log_count": len(logs),
            "format": "json",
            "file_size": Path(output_path).stat().st_size
        }
    
    def _export_csv(self, logs: list, output_path: str, include_timestamp: bool,
                    include_level: bool, include_duration: bool) -> dict:
        """导出为CSV格式"""
        import csv
        
        # 确定列
        columns = []
        if include_timestamp:
            columns.append('timestamp')
        if include_level:
            columns.append('level')
        columns.append('message')
        if include_duration:
            columns.append('duration')
        columns.append('nodeId')
        
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()
            for log in logs:
                row = {
                    'timestamp': log.get('timestamp', ''),
                    'level': log.get('level', ''),
                    'message': log.get('message', ''),
                    'duration': log.get('duration', ''),
                    'nodeId': log.get('nodeId', '')
                }
                writer.writerow(row)
        
        return {
            "output_path": output_path,
            "log_count": len(logs),
            "format": "csv",
            "file_size": Path(output_path).stat().st_size
        }



@register_executor
class ClickTextExecutor(ModuleExecutor):
    """点击文本模块执行器 - 通过屏幕OCR识别实现鼠标点击指定文本"""
    
    @property
    def module_type(self) -> str:
        return "click_text"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        target_text = context.resolve_value(config.get('targetText', ''))
        match_mode = config.get('matchMode', 'contains')  # exact, contains, regex
        click_button = config.get('clickButton', 'left')  # left, right, middle
        click_type = config.get('clickType', 'single')  # single, double
        occurrence = int(config.get('occurrence', 1))  # 第几个匹配项
        search_region = config.get('searchRegion', None)  # 搜索区域 {x, y, width, height}
        wait_timeout = int(config.get('waitTimeout', 10))  # 等待超时（秒）
        result_variable = config.get('resultVariable', '')
        
        if not target_text:
            return ModuleResult(success=False, error="目标文本不能为空")
        
        try:
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None, self._click_text, target_text, match_mode, click_button, 
                click_type, occurrence, search_region, wait_timeout
            )
            
            if result_variable:
                context.set_variable(result_variable, result)
            
            if result.get('found'):
                return ModuleResult(
                    success=True,
                    message=f"已点击文本 \"{target_text}\" 位置: ({result['x']}, {result['y']})",
                    data=result
                )
            else:
                return ModuleResult(success=False, error=f"未找到文本: {target_text}")
        except Exception as e:
            return ModuleResult(success=False, error=f"点击文本失败: {str(e)}")
    
    def _click_text(self, target_text: str, match_mode: str, click_button: str,
                    click_type: str, occurrence: int, search_region: dict, 
                    wait_timeout: int) -> dict:
        """执行OCR识别并点击文本 - 使用 RapidOCR，速度快"""
        import ctypes
        import re
        import numpy as np
        
        try:
            from PIL import ImageGrab
        except ImportError:
            raise ImportError("请安装 Pillow: pip install Pillow")
        
        # 使用 RapidOCR - 比 EasyOCR 快很多
        try:
            from rapidocr_onnxruntime import RapidOCR
            ocr = RapidOCR()
        except ImportError:
            raise ImportError("请安装 rapidocr-onnxruntime: pip install rapidocr-onnxruntime")
        
        start_time = time.time()
        
        while time.time() - start_time < wait_timeout:
            # 截取屏幕
            # 解析搜索区域（支持两点模式和起点+宽高模式）
            region_x, region_y, region_w, region_h = parse_search_region(search_region)
            if region_w > 0 and region_h > 0:
                screenshot = ImageGrab.grab(bbox=(region_x, region_y, region_x + region_w, region_y + region_h))
                offset_x, offset_y = region_x, region_y
            else:
                screenshot = ImageGrab.grab()
                offset_x, offset_y = 0, 0
            
            # 转换为 numpy 数组
            img_array = np.array(screenshot)
            
            # OCR识别
            try:
                result, _ = ocr(img_array)
            except Exception as e:
                print(f"[点击文本] OCR识别失败: {e}")
                time.sleep(0.3)
                continue
            
            if not result:
                time.sleep(0.3)
                continue
            
            # 查找匹配的文本
            matches = []
            for item in result:
                # item 格式: [box, text, confidence]
                # box 格式: [[x1,y1], [x2,y1], [x2,y2], [x1,y2]]
                box, recognized_text, confidence = item
                
                if not recognized_text:
                    continue
                
                # 匹配检查
                is_match = False
                if match_mode == 'exact':
                    is_match = recognized_text == target_text
                elif match_mode == 'contains':
                    is_match = target_text in recognized_text
                elif match_mode == 'regex':
                    try:
                        is_match = bool(re.search(target_text, recognized_text))
                    except:
                        is_match = False
                
                if is_match:
                    x1 = int(min(p[0] for p in box))
                    y1 = int(min(p[1] for p in box))
                    x2 = int(max(p[0] for p in box))
                    y2 = int(max(p[1] for p in box))
                    
                    center_x = (x1 + x2) // 2 + offset_x
                    center_y = (y1 + y2) // 2 + offset_y
                    matches.append({
                        'text': recognized_text,
                        'x': center_x,
                        'y': center_y,
                        'box': [x1 + offset_x, y1 + offset_y, x2 + offset_x, y2 + offset_y],
                        'confidence': confidence
                    })
            
            # 检查是否找到足够的匹配
            if len(matches) >= occurrence:
                match = matches[occurrence - 1]
                
                # 执行点击
                self._perform_click(match['x'], match['y'], click_button, click_type)
                
                return {
                    'found': True,
                    'text': match['text'],
                    'x': match['x'],
                    'y': match['y'],
                    'box': match['box'],
                    'total_matches': len(matches)
                }
            
            # 等待后重试
            time.sleep(0.3)
        
        return {'found': False, 'text': target_text}
    
    def _perform_click(self, x: int, y: int, button: str, click_type: str):
        """执行鼠标点击"""
        import ctypes
        
        user32 = ctypes.windll.user32
        
        # 移动鼠标
        user32.SetCursorPos(x, y)
        time.sleep(0.05)
        
        # 鼠标事件常量
        MOUSEEVENTF_LEFTDOWN = 0x0002
        MOUSEEVENTF_LEFTUP = 0x0004
        MOUSEEVENTF_RIGHTDOWN = 0x0008
        MOUSEEVENTF_RIGHTUP = 0x0010
        MOUSEEVENTF_MIDDLEDOWN = 0x0020
        MOUSEEVENTF_MIDDLEUP = 0x0040
        
        if button == 'left':
            down_event = MOUSEEVENTF_LEFTDOWN
            up_event = MOUSEEVENTF_LEFTUP
        elif button == 'right':
            down_event = MOUSEEVENTF_RIGHTDOWN
            up_event = MOUSEEVENTF_RIGHTUP
        else:
            down_event = MOUSEEVENTF_MIDDLEDOWN
            up_event = MOUSEEVENTF_MIDDLEUP
        
        # 执行点击
        clicks = 2 if click_type == 'double' else 1
        for _ in range(clicks):
            user32.mouse_event(down_event, 0, 0, 0, 0)
            time.sleep(0.02)
            user32.mouse_event(up_event, 0, 0, 0, 0)
            if click_type == 'double':
                time.sleep(0.05)


@register_executor
class HoverImageExecutor(ModuleExecutor):
    """鼠标悬停在图像上模块执行器 - 在屏幕上查找指定图像并将鼠标悬停在上面"""

    @property
    def module_type(self) -> str:
        return "hover_image"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import ctypes

        image_path = context.resolve_value(config.get("imagePath", ""))
        confidence = to_float(config.get("confidence", 0.8), 0.8, context)
        hover_position = context.resolve_value(config.get("hoverPosition", "center"))
        hover_duration = to_float(config.get("hoverDuration", 0.5), 0.5, context)  # 悬停时长（秒）
        wait_timeout = to_int(config.get("waitTimeout", 10), 10, context)
        result_variable = config.get("resultVariable", "")
        search_region = config.get("searchRegion", None)  # 搜索区域

        if not image_path:
            return ModuleResult(success=False, error="图像路径不能为空")

        if not Path(image_path).exists():
            return ModuleResult(success=False, error=f"图像文件不存在: {image_path}")

        try:
            import cv2
            import numpy as np
            
            # 设置 DPI 感知
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except:
                    pass
            
            # 读取模板图像
            template = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_COLOR)
            if template is None:
                return ModuleResult(success=False, error="无法读取图像文件")

            template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
            h, w = template_gray.shape

            # 解析搜索区域（支持两点模式和起点+宽高模式）
            region_x, region_y, region_w, region_h = parse_search_region(search_region)
            use_region = region_w > 0 and region_h > 0

            # 获取虚拟屏幕尺寸
            SM_XVIRTUALSCREEN = 76
            SM_YVIRTUALSCREEN = 77
            SM_CXVIRTUALSCREEN = 78
            SM_CYVIRTUALSCREEN = 79
            
            virtual_left = ctypes.windll.user32.GetSystemMetrics(SM_XVIRTUALSCREEN)
            virtual_top = ctypes.windll.user32.GetSystemMetrics(SM_YVIRTUALSCREEN)
            virtual_width = ctypes.windll.user32.GetSystemMetrics(SM_CXVIRTUALSCREEN)
            virtual_height = ctypes.windll.user32.GetSystemMetrics(SM_CYVIRTUALSCREEN)

            start_time = time.time()
            found = False
            hover_x, hover_y = 0, 0
            best_confidence = 0

            while time.time() - start_time < wait_timeout:
                # 截取屏幕
                if use_region:
                    # 使用指定区域截图
                    from PIL import ImageGrab
                    screenshot_pil = ImageGrab.grab(bbox=(region_x, region_y, region_x + region_w, region_y + region_h))
                    screen = cv2.cvtColor(np.array(screenshot_pil), cv2.COLOR_RGB2BGR)
                    offset_x, offset_y = region_x, region_y
                else:
                    try:
                        import win32gui
                        import win32ui
                        import win32con
                        
                        hdesktop = win32gui.GetDesktopWindow()
                        desktop_dc = win32gui.GetWindowDC(hdesktop)
                        img_dc = win32ui.CreateDCFromHandle(desktop_dc)
                        mem_dc = img_dc.CreateCompatibleDC()
                        
                        screenshot = win32ui.CreateBitmap()
                        screenshot.CreateCompatibleBitmap(img_dc, virtual_width, virtual_height)
                        mem_dc.SelectObject(screenshot)
                        mem_dc.BitBlt((0, 0), (virtual_width, virtual_height), img_dc, 
                                      (virtual_left, virtual_top), win32con.SRCCOPY)
                        
                        bmpinfo = screenshot.GetInfo()
                        bmpstr = screenshot.GetBitmapBits(True)
                        screen = np.frombuffer(bmpstr, dtype=np.uint8).reshape(
                            (bmpinfo['bmHeight'], bmpinfo['bmWidth'], 4))
                        screen = cv2.cvtColor(screen, cv2.COLOR_BGRA2BGR)
                        
                        mem_dc.DeleteDC()
                        win32gui.DeleteObject(screenshot.GetHandle())
                        win32gui.ReleaseDC(hdesktop, desktop_dc)
                        offset_x, offset_y = virtual_left, virtual_top
                        
                    except ImportError:
                        from PIL import ImageGrab
                        screenshot_pil = ImageGrab.grab(all_screens=True)
                        screen = cv2.cvtColor(np.array(screenshot_pil), cv2.COLOR_RGB2BGR)
                        offset_x, offset_y = 0, 0

                screen_gray = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)
                result = cv2.matchTemplate(screen_gray, template_gray, cv2.TM_CCOEFF_NORMED)
                min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

                if max_val >= confidence:
                    img_left = offset_x + max_loc[0]
                    img_top = offset_y + max_loc[1]
                    img_right = img_left + w
                    img_bottom = img_top + h
                    
                    hover_x, hover_y = self._calculate_hover_position(
                        hover_position, img_left, img_top, img_right, img_bottom, w, h
                    )
                    best_confidence = max_val
                    found = True
                    break

                await asyncio.sleep(0.3)

            if not found:
                return ModuleResult(
                    success=False, 
                    error=f"在 {wait_timeout} 秒内未找到匹配的图像（最高匹配度: {best_confidence:.2%}）"
                )

            # 移动鼠标到目标位置
            user32 = ctypes.windll.user32
            user32.SetCursorPos(int(hover_x), int(hover_y))
            
            # 悬停指定时长
            await asyncio.sleep(hover_duration)

            result_data = {"x": hover_x, "y": hover_y, "confidence": best_confidence, "position": hover_position}
            if result_variable:
                context.set_variable(result_variable, result_data)

            position_name = self._get_position_name(hover_position)
            return ModuleResult(
                success=True, 
                message=f"已在图像{position_name} ({hover_x}, {hover_y}) 悬停 {hover_duration}ms，匹配度: {best_confidence:.2%}",
                data=result_data
            )

        except ImportError as e:
            missing = str(e).split("'")[1] if "'" in str(e) else "opencv-python/Pillow"
            return ModuleResult(success=False, error=f"需要安装依赖库: pip install {missing}")
        except Exception as e:
            return ModuleResult(success=False, error=f"悬停图像失败: {str(e)}")
    
    def _calculate_hover_position(self, position: str, left: int, top: int, right: int, bottom: int, w: int, h: int) -> tuple:
        """根据悬停位置计算实际坐标"""
        if position == "center":
            return (left + w // 2, top + h // 2)
        elif position == "top-left":
            return (left + 5, top + 5)
        elif position == "top-right":
            return (right - 5, top + 5)
        elif position == "bottom-left":
            return (left + 5, bottom - 5)
        elif position == "bottom-right":
            return (right - 5, bottom - 5)
        elif position == "top":
            return (left + w // 2, top + 5)
        elif position == "bottom":
            return (left + w // 2, bottom - 5)
        elif position == "left":
            return (left + 5, top + h // 2)
        elif position == "right":
            return (right - 5, top + h // 2)
        elif position == "random":
            margin = 5
            rand_x = random.randint(left + margin, right - margin)
            rand_y = random.randint(top + margin, bottom - margin)
            return (rand_x, rand_y)
        else:
            return (left + w // 2, top + h // 2)
    
    def _get_position_name(self, position: str) -> str:
        """获取位置的中文名称"""
        names = {
            "center": "中心",
            "top-left": "左上角",
            "top-right": "右上角",
            "bottom-left": "左下角",
            "bottom-right": "右下角",
            "top": "顶部",
            "bottom": "底部",
            "left": "左侧",
            "right": "右侧",
            "random": "随机位置"
        }
        return names.get(position, "中心")


@register_executor
class HoverTextExecutor(ModuleExecutor):
    """鼠标悬停在文本上模块执行器 - 通过屏幕OCR识别实现鼠标悬停在指定文本上"""
    
    @property
    def module_type(self) -> str:
        return "hover_text"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        target_text = context.resolve_value(config.get('targetText', ''))
        match_mode = config.get('matchMode', 'contains')  # exact, contains, regex
        hover_duration = to_int(config.get('hoverDuration', 500), 500, context)  # 悬停时长（毫秒）
        occurrence = int(config.get('occurrence', 1))  # 第几个匹配项
        search_region = config.get('searchRegion', None)  # 搜索区域
        wait_timeout = int(config.get('waitTimeout', 10))  # 等待超时（秒）
        result_variable = config.get('resultVariable', '')
        
        if not target_text:
            return ModuleResult(success=False, error="目标文本不能为空")
        
        try:
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None, self._hover_text, target_text, match_mode, hover_duration,
                occurrence, search_region, wait_timeout
            )
            
            if result_variable:
                context.set_variable(result_variable, result)
            
            if result.get('found'):
                return ModuleResult(
                    success=True,
                    message=f"已在文本 \"{target_text}\" 位置 ({result['x']}, {result['y']}) 悬停 {hover_duration}ms",
                    data=result
                )
            else:
                return ModuleResult(success=False, error=f"未找到文本: {target_text}")
        except Exception as e:
            return ModuleResult(success=False, error=f"悬停文本失败: {str(e)}")
    
    def _hover_text(self, target_text: str, match_mode: str, hover_duration: int,
                    occurrence: int, search_region: dict, wait_timeout: int) -> dict:
        """执行OCR识别并悬停在文本上 - 使用 RapidOCR，速度快"""
        import ctypes
        import re
        import numpy as np
        
        # 调试日志
        print(f"[悬停文本] 目标文本: '{target_text}', 匹配模式: {match_mode}")
        print(f"[悬停文本] search_region 原始值: {search_region}")
        
        try:
            from PIL import ImageGrab
        except ImportError:
            raise ImportError("请安装 Pillow: pip install Pillow")
        
        # 使用 RapidOCR - 比 EasyOCR 快很多
        try:
            from rapidocr_onnxruntime import RapidOCR
            ocr = RapidOCR()
        except ImportError:
            raise ImportError("请安装 rapidocr-onnxruntime: pip install rapidocr-onnxruntime")
        
        start_time = time.time()
        first_loop = True
        
        while time.time() - start_time < wait_timeout:
            # 截取屏幕
            # 解析搜索区域（支持两点模式和起点+宽高模式）
            region_x, region_y, region_w, region_h = parse_search_region(search_region)
            
            if first_loop:
                print(f"[悬停文本] 解析后区域: x={region_x}, y={region_y}, w={region_w}, h={region_h}")
                first_loop = False
            
            if region_w > 0 and region_h > 0:
                bbox = (region_x, region_y, region_x + region_w, region_y + region_h)
                print(f"[悬停文本] 截图区域 bbox: {bbox}")
                screenshot = ImageGrab.grab(bbox=bbox)
                offset_x, offset_y = region_x, region_y
            else:
                screenshot = ImageGrab.grab()
                offset_x, offset_y = 0, 0
            
            # 转换为 numpy 数组
            img_array = np.array(screenshot)
            print(f"[悬停文本] 截图尺寸: {img_array.shape}")
            
            # OCR识别
            try:
                result, _ = ocr(img_array)
            except Exception as e:
                print(f"[悬停文本] OCR识别失败: {e}")
                time.sleep(0.3)
                continue
            
            if not result:
                print(f"[悬停文本] OCR未识别到任何文本")
                time.sleep(0.3)
                continue
            
            # 打印识别到的所有文本
            print(f"[悬停文本] OCR识别到 {len(result)} 个文本区域:")
            for idx, item in enumerate(result):
                _, text, conf = item
                try:
                    conf_val = float(conf) if conf is not None else 0
                    print(f"  [{idx}] '{text}' (置信度: {conf_val:.2f})")
                except (ValueError, TypeError):
                    print(f"  [{idx}] '{text}' (置信度: {conf})")
            
            # 查找匹配的文本
            matches = []
            for item in result:
                # item 格式: [box, text, confidence]
                # box 格式: [[x1,y1], [x2,y1], [x2,y2], [x1,y2]]
                box, recognized_text, confidence = item
                
                if not recognized_text:
                    continue
                
                is_match = False
                if match_mode == 'exact':
                    is_match = recognized_text == target_text
                elif match_mode == 'contains':
                    is_match = target_text in recognized_text
                elif match_mode == 'regex':
                    try:
                        is_match = bool(re.search(target_text, recognized_text))
                    except:
                        is_match = False
                
                if is_match:
                    x1 = int(min(p[0] for p in box))
                    y1 = int(min(p[1] for p in box))
                    x2 = int(max(p[0] for p in box))
                    y2 = int(max(p[1] for p in box))
                    
                    center_x = (x1 + x2) // 2 + offset_x
                    center_y = (y1 + y2) // 2 + offset_y
                    matches.append({
                        'text': recognized_text,
                        'x': center_x,
                        'y': center_y,
                        'box': [x1 + offset_x, y1 + offset_y, x2 + offset_x, y2 + offset_y],
                        'confidence': confidence
                    })
            
            if len(matches) >= occurrence:
                match = matches[occurrence - 1]
                
                # 移动鼠标到目标位置并悬停
                user32 = ctypes.windll.user32
                user32.SetCursorPos(match['x'], match['y'])
                time.sleep(hover_duration / 1000)
                
                return {
                    'found': True,
                    'text': match['text'],
                    'x': match['x'],
                    'y': match['y'],
                    'box': match['box'],
                    'total_matches': len(matches)
                }
            
            time.sleep(0.3)
        
        return {'found': False, 'text': target_text}




@register_executor
class ShareFolderExecutor(ModuleExecutor):
    """文件夹网络共享模块执行器 - 将指定文件夹通过HTTP共享到局域网"""

    @property
    def module_type(self) -> str:
        return "share_folder"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        from app.services.file_share import start_file_share, get_local_ip
        
        folder_path = context.resolve_value(config.get("folderPath", ""))
        port = to_int(config.get("port", 8080), 8080, context)
        share_name = context.resolve_value(config.get("shareName", "")) or "共享文件夹"
        result_variable = config.get("resultVariable", "share_url")
        allow_write = config.get("allowWrite", True)  # 默认允许写操作
        
        if not folder_path:
            return ModuleResult(success=False, error="文件夹路径不能为空")
        
        folder = Path(folder_path)
        if not folder.exists():
            return ModuleResult(success=False, error=f"文件夹不存在: {folder_path}")
        
        if not folder.is_dir():
            return ModuleResult(success=False, error=f"路径不是文件夹: {folder_path}")
        
        try:
            result = start_file_share(
                path=str(folder.resolve()),
                port=port,
                share_type='folder',
                name=share_name,
                allow_write=allow_write
            )
            
            if not result['success']:
                return ModuleResult(success=False, error=result.get('error', '启动共享服务失败'))
            
            share_url = result['url']
            local_ip = result['ip']
            
            if result_variable:
                context.set_variable(result_variable, share_url)
            
            write_mode = "可上传/删除" if allow_write else "仅下载"
            message = f"📂 文件夹共享已启动！\n" \
                      f"共享名称: {share_name}\n" \
                      f"共享路径: {folder_path}\n" \
                      f"访问地址: {share_url}\n" \
                      f"权限模式: {write_mode}\n" \
                      f"💡 同局域网的设备可以使用浏览器访问上述地址来浏览和下载文件"
            
            return ModuleResult(
                success=True,
                message=message,
                data={
                    'url': share_url,
                    'ip': local_ip,
                    'port': port,
                    'path': str(folder.resolve()),
                    'name': share_name,
                    'allowWrite': allow_write
                }
            )
            
        except Exception as e:
            return ModuleResult(success=False, error=f"启动文件夹共享失败: {str(e)}")


@register_executor
class ShareFileExecutor(ModuleExecutor):
    """文件网络共享模块执行器 - 将指定文件通过HTTP共享到局域网"""

    @property
    def module_type(self) -> str:
        return "share_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        from app.services.file_share import start_file_share, get_local_ip
        
        file_path = context.resolve_value(config.get("filePath", ""))
        port = to_int(config.get("port", 8080), 8080, context)
        result_variable = config.get("resultVariable", "share_url")
        
        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")
        
        file = Path(file_path)
        if not file.exists():
            return ModuleResult(success=False, error=f"文件不存在: {file_path}")
        
        if not file.is_file():
            return ModuleResult(success=False, error=f"路径不是文件: {file_path}")
        
        try:
            result = start_file_share(
                path=str(file.resolve()),
                port=port,
                share_type='file',
                name=file.name
            )
            
            if not result['success']:
                return ModuleResult(success=False, error=result.get('error', '启动共享服务失败'))
            
            share_url = result['url']
            local_ip = result['ip']
            file_size = file.stat().st_size
            
            # 格式化文件大小
            size_str = self._format_size(file_size)
            
            if result_variable:
                context.set_variable(result_variable, share_url)
            
            message = f"📄 文件共享已启动！\n" \
                      f"文件名: {file.name}\n" \
                      f"文件大小: {size_str}\n" \
                      f"访问地址: {share_url}\n" \
                      f"💡 同局域网的设备可以使用浏览器访问上述地址来下载此文件"
            
            return ModuleResult(
                success=True,
                message=message,
                data={
                    'url': share_url,
                    'ip': local_ip,
                    'port': port,
                    'path': str(file.resolve()),
                    'name': file.name,
                    'size': file_size
                }
            )
            
        except Exception as e:
            return ModuleResult(success=False, error=f"启动文件共享失败: {str(e)}")
    
    def _format_size(self, size: int) -> str:
        """格式化文件大小"""
        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if size < 1024:
                return f"{size:.1f} {unit}" if unit != 'B' else f"{size} {unit}"
            size /= 1024
        return f"{size:.1f} PB"


@register_executor
class StopShareExecutor(ModuleExecutor):
    """停止网络共享模块执行器"""

    @property
    def module_type(self) -> str:
        return "stop_share"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        from app.services.file_share import stop_file_share
        
        port = to_int(config.get("port", 8080), 8080, context)
        
        try:
            result = stop_file_share(port)
            
            if not result['success']:
                return ModuleResult(success=False, error=result.get('error', '停止共享服务失败'))
            
            return ModuleResult(
                success=True,
                message=f"端口 {port} 的共享服务已停止",
                data={'port': port}
            )
            
        except Exception as e:
            return ModuleResult(success=False, error=f"停止共享服务失败: {str(e)}")
