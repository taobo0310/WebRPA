"""数据收集器 - 使用Polars管理和导出数据"""
import os
from pathlib import Path
from typing import Any, Optional
from datetime import datetime

import polars as pl


class DataCollector:
    """数据收集器"""
    
    def __init__(self):
        self.data: list[dict[str, Any]] = []
        self.columns: list[str] = []
        self._current_row: dict[str, Any] = {}
    
    def add_value(self, column: str, value: Any):
        """添加单个值到当前行"""
        if column not in self.columns:
            self.columns.append(column)
        self._current_row[column] = value
    
    def add_row(self, row: dict[str, Any]):
        """添加一行数据"""
        for key in row:
            if key not in self.columns:
                self.columns.append(key)
        self.data.append(row)
    
    def commit_row(self):
        """提交当前行"""
        if self._current_row:
            self.add_row(self._current_row.copy())
            self._current_row = {}
    
    def clear(self):
        """清空数据"""
        self.data = []
        self.columns = []
        self._current_row = {}
    
    def to_dataframe(self) -> pl.DataFrame:
        """转换为Polars DataFrame"""
        if not self.data:
            return pl.DataFrame()
        
        # 确保所有行都有相同的列
        normalized_data = []
        for row in self.data:
            normalized_row = {col: row.get(col, None) for col in self.columns}
            normalized_data.append(normalized_row)
        
        return pl.DataFrame(normalized_data)
    
    def to_excel(self, filepath: str, sheet_name: str = '数据') -> str:
        """导出为Excel文件（带样式）
        
        Args:
            filepath: 文件路径
            sheet_name: Sheet名称，默认为'数据'
        """
        df = self.to_dataframe()
        
        # 确保目录存在
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        
        # 导出Excel（带样式）
        _write_styled_excel(df, filepath, sheet_name)
        
        return filepath
    
    def to_csv(self, filepath: str) -> str:
        """导出为CSV文件"""
        df = self.to_dataframe()
        
        # 确保目录存在
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        
        # 导出CSV
        df.write_csv(filepath)
        
        return filepath
    
    @property
    def row_count(self) -> int:
        """获取行数"""
        return len(self.data)
    
    @property
    def column_count(self) -> int:
        """获取列数"""
        return len(self.columns)


def _write_styled_excel(df: pl.DataFrame, filepath: str, sheet_name: str = '数据'):
    """写入带样式的Excel文件
    
    Args:
        df: Polars DataFrame
        filepath: 文件路径
        sheet_name: Sheet名称，默认为'数据'
    """
    try:
        from xlsxwriter import Workbook
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # 检查文件是否已存在
        file_exists = Path(filepath).exists()
        
        if file_exists:
            # 文件存在，使用openpyxl追加或更新Sheet
            try:
                wb = openpyxl.load_workbook(filepath)
                
                # 如果Sheet已存在，删除它
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                # 创建新的Sheet
                ws = wb.create_sheet(sheet_name)
                
                # 写入数据（使用pandas DataFrame作为中间格式）
                import pandas as pd
                pandas_df = df.to_pandas()
                
                # 写入表头
                for col_idx, col_name in enumerate(pandas_df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    # 设置表头样式
                    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin', color='2F5496'),
                        right=openpyxl.styles.Side(style='thin', color='2F5496'),
                        top=openpyxl.styles.Side(style='thin', color='2F5496'),
                        bottom=openpyxl.styles.Side(style='thin', color='2F5496')
                    )
                
                # 写入数据行
                for row_idx, row_data in enumerate(pandas_df.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # 设置数据样式
                        if row_idx % 2 == 0:
                            cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                            right=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                            top=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                            bottom=openpyxl.styles.Side(style='thin', color='D9D9D9')
                        )
                
                # 自动调整列宽
                for col_idx, col_name in enumerate(pandas_df.columns, 1):
                    max_len = len(str(col_name))
                    for value in pandas_df.iloc[:, col_idx-1]:
                        cell_len = len(str(value)) if value is not None else 0
                        max_len = max(max_len, cell_len)
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)
                
                # 冻结首行
                ws.freeze_panes = 'A2'
                
                # 保存工作簿
                wb.save(filepath)
                
            except Exception as e:
                # 如果使用openpyxl失败，回退到xlsxwriter（会覆盖整个文件）
                print(f"[警告] 使用openpyxl追加Sheet失败: {e}，将使用xlsxwriter创建新文件")
                file_exists = False
        
        if not file_exists:
            # 文件不存在，使用xlsxwriter创建新文件
            # 创建工作簿
            workbook = Workbook(filepath)
            worksheet = workbook.add_worksheet(sheet_name)
            
            # 定义样式
            header_format = workbook.add_format({
                'bold': True,
                'font_size': 11,
                'font_color': 'white',
                'bg_color': '#4472C4',
                'border': 1,
                'border_color': '#2F5496',
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True,
            })
            
            cell_format = workbook.add_format({
                'font_size': 10,
                'border': 1,
                'border_color': '#D9D9D9',
                'align': 'left',
                'valign': 'vcenter',
            })
            
            alt_cell_format = workbook.add_format({
                'font_size': 10,
                'border': 1,
                'border_color': '#D9D9D9',
                'bg_color': '#F2F2F2',
                'align': 'left',
                'valign': 'vcenter',
            })
            
            # 写入表头
            columns = df.columns
            for col_idx, col_name in enumerate(columns):
                worksheet.write(0, col_idx, col_name, header_format)
            
            # 写入数据
            for row_idx, row in enumerate(df.iter_rows()):
                row_format = alt_cell_format if row_idx % 2 == 1 else cell_format
                for col_idx, value in enumerate(row):
                    worksheet.write(row_idx + 1, col_idx, value, row_format)
            
            # 自动调整列宽
            for col_idx, col_name in enumerate(columns):
                # 计算列宽（取表头和数据中最长的）
                max_len = len(str(col_name))
                for row in df.iter_rows():
                    cell_len = len(str(row[col_idx])) if row[col_idx] is not None else 0
                    max_len = max(max_len, cell_len)
                # 设置列宽（加一些padding）
                worksheet.set_column(col_idx, col_idx, min(max_len + 4, 50))
            
            # 设置行高
            worksheet.set_row(0, 25)  # 表头行高
            
            # 冻结首行
            worksheet.freeze_panes(1, 0)
            
            workbook.close()
        
    except ImportError as e:
        # 如果没有xlsxwriter或openpyxl，使用polars默认导出
        print(f"[警告] 缺少必要的库: {e}，使用polars默认导出")
        df.write_excel(filepath)


class DataExporter:
    """数据导出器"""
    
    def __init__(self, output_dir: str = "./data"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(self, data: list[dict], filename: Optional[str] = None) -> str:
        """导出数据到Excel（带样式）"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        if not data:
            # 创建空文件
            pl.DataFrame().write_excel(str(filepath))
        else:
            df = pl.DataFrame(data)
            _write_styled_excel(df, str(filepath))
        
        return str(filepath)
    
    def export_to_csv(self, data: list[dict], filename: Optional[str] = None) -> str:
        """导出数据到CSV"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        if not data:
            # 创建空文件
            pl.DataFrame().write_csv(str(filepath))
        else:
            df = pl.DataFrame(data)
            df.write_csv(str(filepath))
        
        return str(filepath)
