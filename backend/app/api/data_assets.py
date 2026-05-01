"""Excel文件资源API - 处理Excel文件上传和读取"""
import os
import uuid
import shutil
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, UploadFile, File, HTTPException
from pydantic import BaseModel
import openpyxl
import xlrd

router = APIRouter(prefix="/api/data-assets", tags=["data-assets"])

# 存储上传的文件信息
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'uploads', 'excel')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 内存中存储文件元数据
data_assets: dict[str, dict] = {}


class CreateFolderRequest(BaseModel):
    name: str
    parentPath: Optional[str] = None


class RenameFolderRequest(BaseModel):
    oldPath: str
    newName: str


class MoveAssetRequest(BaseModel):
    assetId: str
    targetFolder: Optional[str] = None


class DeleteFolderRequest(BaseModel):
    folderPath: str


class ReadExcelRequest(BaseModel):
    fileId: str
    sheetName: Optional[str] = None
    readMode: str  # 'cell', 'row', 'column', 'range'
    cellAddress: Optional[str] = None  # 如 'A1'
    rowIndex: Optional[int] = None  # 行号，从1开始
    columnIndex: Optional[int] = None  # 列号，从1开始，或列字母如'A'
    startCell: Optional[str] = None  # 范围起始，如 'A1'
    endCell: Optional[str] = None  # 范围结束，如 'C10'


def _get_full_path(relative_path: Optional[str] = None) -> str:
    """获取完整路径"""
    if not relative_path:
        return UPLOAD_DIR
    return os.path.join(UPLOAD_DIR, relative_path.replace('/', os.sep))


def _get_relative_path(full_path: str) -> str:
    """获取相对路径"""
    rel = os.path.relpath(full_path, UPLOAD_DIR)
    if rel == '.':
        return ''
    return rel.replace(os.sep, '/')


def _load_existing_files():
    """启动时扫描excel文件夹,加载已存在的Excel文件"""
    if not os.path.exists(UPLOAD_DIR):
        return
    
    for root, dirs, files in os.walk(UPLOAD_DIR):
        for filename in files:
            if not filename.lower().endswith(('.xlsx', '.xls')):
                continue
            
            file_path = os.path.join(root, filename)
            
            try:
                # 从文件名提取ID (格式: uuid.xlsx 或任意名称.xlsx)
                name_without_ext = os.path.splitext(filename)[0]
                ext = os.path.splitext(filename)[1]
                
                # 如果文件名是UUID格式,使用它作为ID;否则生成新的UUID
                if len(name_without_ext) == 36 and name_without_ext.count('-') == 4:
                    file_id = name_without_ext
                else:
                    # 为非UUID格式的文件生成新ID,但保留原始文件名
                    file_id = str(uuid.uuid4())
                
                # 读取工作表名称
                is_xls = ext.lower() == '.xls'
                if is_xls:
                    wb = xlrd.open_workbook(file_path)
                    sheet_names = wb.sheet_names()
                else:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                
                # 获取文件信息
                file_stat = os.stat(file_path)
                
                # 获取相对于UPLOAD_DIR的文件夹路径
                folder_path = _get_relative_path(root)
                
                # 存储元数据
                asset = {
                    'id': file_id,
                    'name': filename,  # 保存实际文件名
                    'originalName': filename,  # 原始文件名
                    'size': file_stat.st_size,
                    'uploadedAt': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                    'sheetNames': sheet_names,
                    'path': file_path,
                    'folder': folder_path,
                }
                data_assets[file_id] = asset
                print(f"[DataAssets] 已加载Excel文件: {filename} (文件夹: {folder_path or '根目录'})")
            except Exception as e:
                print(f"[DataAssets] 加载Excel文件失败 {filename}: {str(e)}")
                continue


# 启动时加载已存在的文件
_load_existing_files()


@router.post("/upload")
async def upload_excel(file: UploadFile = File(...), folder: Optional[str] = None):
    """上传Excel文件"""
    # 检查文件类型
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持Excel文件 (.xlsx, .xls)")
    
    # 生成唯一ID和文件名
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(file.filename)[1]
    saved_name = f"{file_id}{ext}"
    
    # 确定保存路径
    target_dir = _get_full_path(folder)
    os.makedirs(target_dir, exist_ok=True)
    file_path = os.path.join(target_dir, saved_name)
    
    # 保存文件
    content = await file.read()
    with open(file_path, 'wb') as f:
        f.write(content)
    
    # 读取工作表名称（根据文件格式选择不同的库）
    try:
        is_xls = ext.lower() == '.xls'
        if is_xls:
            # 使用xlrd读取旧版.xls文件
            wb = xlrd.open_workbook(file_path)
            sheet_names = wb.sheet_names()
        else:
            # 使用openpyxl读取.xlsx文件
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
    except Exception as e:
        os.remove(file_path)
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件: {str(e)}")
    
    # 存储元数据
    asset = {
        'id': file_id,
        'name': saved_name,
        'originalName': file.filename,
        'size': len(content),
        'uploadedAt': datetime.now().isoformat(),
        'sheetNames': sheet_names,
        'path': file_path,
        'folder': folder or '',
    }
    data_assets[file_id] = asset
    
    return {
        'asset': {
            'id': asset['id'],
            'name': asset['name'],
            'originalName': asset['originalName'],
            'size': asset['size'],
            'uploadedAt': asset['uploadedAt'],
            'sheetNames': asset['sheetNames'],
            'folder': asset['folder'],
        }
    }


@router.post("/upload-batch")
async def upload_excel_batch(files: List[UploadFile] = File(...), folder: Optional[str] = None):
    """批量上传Excel文件"""
    results = []
    errors = []
    
    for file in files:
        try:
            result = await upload_excel(file, folder)
            results.append(result)
        except Exception as e:
            errors.append({'filename': file.filename, 'error': str(e)})
    
    return {
        'success': results,
        'errors': errors,
        'total': len(files),
        'successCount': len(results),
        'errorCount': len(errors),
    }


@router.get("")
async def list_assets(folder: Optional[str] = None):
    """获取所有Excel文件资源"""
    # 如果指定了文件夹,只返回该文件夹下的文件
    if folder is not None:
        filtered = [
            {
                'id': a['id'],
                'name': a['name'],
                'originalName': a['originalName'],
                'size': a['size'],
                'uploadedAt': a['uploadedAt'],
                'sheetNames': a['sheetNames'],
                'folder': a['folder'],
            }
            for a in data_assets.values()
            if a['folder'] == folder
        ]
    else:
        # 返回所有文件
        filtered = [
            {
                'id': a['id'],
                'name': a['name'],
                'originalName': a['originalName'],
                'size': a['size'],
                'uploadedAt': a['uploadedAt'],
                'sheetNames': a['sheetNames'],
                'folder': a['folder'],
            }
            for a in data_assets.values()
        ]
    
    return filtered


@router.get("/folders")
async def list_folders():
    """获取所有文件夹列表"""
    folders = set()
    
    # 从已加载的文件中提取文件夹
    for asset in data_assets.values():
        if asset['folder']:
            folders.add(asset['folder'])
    
    # 扫描文件系统中的文件夹
    for root, dirs, files in os.walk(UPLOAD_DIR):
        for dir_name in dirs:
            folder_path = os.path.join(root, dir_name)
            rel_path = _get_relative_path(folder_path)
            folders.add(rel_path)
    
    return sorted(list(folders))


@router.post("/folders")
async def create_folder(request: CreateFolderRequest):
    """创建文件夹"""
    # 验证文件夹名称
    if not request.name or '/' in request.name or '\\' in request.name:
        raise HTTPException(status_code=400, detail="文件夹名称无效")
    
    # 构建完整路径
    if request.parentPath:
        folder_path = os.path.join(request.parentPath, request.name)
    else:
        folder_path = request.name
    
    full_path = _get_full_path(folder_path)
    
    # 检查是否已存在
    if os.path.exists(full_path):
        raise HTTPException(status_code=400, detail="文件夹已存在")
    
    # 创建文件夹
    try:
        os.makedirs(full_path, exist_ok=True)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建文件夹失败: {str(e)}")
    
    return {'success': True, 'path': folder_path}


@router.put("/folders/rename")
async def rename_folder(request: RenameFolderRequest):
    """重命名文件夹"""
    old_full_path = _get_full_path(request.oldPath)
    
    if not os.path.exists(old_full_path):
        raise HTTPException(status_code=404, detail="文件夹不存在")
    
    # 构建新路径
    parent_path = os.path.dirname(old_full_path)
    new_full_path = os.path.join(parent_path, request.newName)
    
    if os.path.exists(new_full_path):
        raise HTTPException(status_code=400, detail="目标文件夹已存在")
    
    try:
        # 重命名文件夹
        os.rename(old_full_path, new_full_path)
        
        # 更新所有该文件夹下的文件元数据
        new_rel_path = _get_relative_path(new_full_path)
        for asset in data_assets.values():
            if asset['folder'].startswith(request.oldPath):
                # 更新文件夹路径
                asset['folder'] = asset['folder'].replace(request.oldPath, new_rel_path, 1)
                # 更新文件路径
                asset['path'] = asset['path'].replace(old_full_path, new_full_path, 1)
        
        return {'success': True, 'newPath': new_rel_path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"重命名失败: {str(e)}")


@router.delete("/folders")
async def delete_folder(request: DeleteFolderRequest):
    """删除文件夹及其所有内容"""
    full_path = _get_full_path(request.folderPath)
    
    if not os.path.exists(full_path):
        raise HTTPException(status_code=404, detail="文件夹不存在")
    
    try:
        # 删除文件夹
        shutil.rmtree(full_path)
        
        # 删除该文件夹下的所有文件元数据
        to_delete = [
            asset_id for asset_id, asset in data_assets.items()
            if asset['folder'].startswith(request.folderPath)
        ]
        for asset_id in to_delete:
            del data_assets[asset_id]
        
        return {'success': True, 'deletedCount': len(to_delete)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.put("/move")
async def move_asset(request: MoveAssetRequest):
    """移动Excel文件到指定文件夹"""
    if request.assetId not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[request.assetId]
    old_path = asset['path']
    
    # 构建新路径
    target_dir = _get_full_path(request.targetFolder)
    os.makedirs(target_dir, exist_ok=True)
    new_path = os.path.join(target_dir, asset['name'])
    
    try:
        # 移动文件
        shutil.move(old_path, new_path)
        
        # 更新元数据
        asset['path'] = new_path
        asset['folder'] = request.targetFolder or ''
        
        return {'success': True, 'newFolder': asset['folder']}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"移动失败: {str(e)}")


@router.delete("/{file_id}")
async def delete_asset(file_id: str):
    """删除Excel文件资源"""
    if file_id not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[file_id]
    
    # 删除文件
    if os.path.exists(asset['path']):
        os.remove(asset['path'])
    
    # 删除元数据
    del data_assets[file_id]
    
    return {'message': '删除成功'}


@router.put("/{file_id}/rename")
async def rename_asset(file_id: str, newName: str):
    """重命名Excel文件"""
    if file_id not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[file_id]
    old_path = asset['path']
    
    # 验证新文件名
    if not newName or '/' in newName or '\\' in newName:
        raise HTTPException(status_code=400, detail="文件名无效")
    
    # 构建新路径
    folder_path = os.path.dirname(old_path)
    new_path = os.path.join(folder_path, newName)
    
    # 检查是否已存在
    if os.path.exists(new_path) and new_path != old_path:
        raise HTTPException(status_code=400, detail="文件名已存在")
    
    try:
        # 重命名文件
        os.rename(old_path, new_path)
        
        # 更新元数据
        asset['path'] = new_path
        asset['name'] = newName
        asset['originalName'] = newName
        
        return {'success': True, 'asset': {
            'id': asset['id'],
            'name': asset['name'],
            'originalName': asset['originalName'],
            'size': asset['size'],
            'uploadedAt': asset['uploadedAt'],
            'folder': asset['folder'],
            'sheetNames': asset['sheetNames'],
        }}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"重命名失败: {str(e)}")


@router.post("/read")
async def read_excel(request: ReadExcelRequest):
    """读取Excel数据"""
    if request.fileId not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[request.fileId]
    file_path = asset['path']
    is_xls = file_path.lower().endswith('.xls')
    
    try:
        if is_xls:
            # 使用xlrd读取.xls文件
            return await _read_excel_xls(file_path, request)
        else:
            # 使用openpyxl读取.xlsx文件
            return await _read_excel_xlsx(file_path, request)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"读取Excel失败: {str(e)}")


async def _read_excel_xlsx(file_path: str, request: ReadExcelRequest):
    """使用openpyxl读取.xlsx文件"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # 选择工作表
    if request.sheetName:
        if request.sheetName not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"工作表 '{request.sheetName}' 不存在")
        ws = wb[request.sheetName]
    else:
        ws = wb.active
    
    result = None
    result_type = 'unknown'
    
    if request.readMode == 'cell':
        if not request.cellAddress:
            raise HTTPException(status_code=400, detail="单元格模式需要指定cellAddress")
        cell = ws[request.cellAddress]
        result = cell.value if cell.value is not None else ''
        result_type = 'cell'
    
    elif request.readMode == 'row':
        if request.rowIndex is None:
            raise HTTPException(status_code=400, detail="行模式需要指定rowIndex")
        row_data = []
        for cell in ws[request.rowIndex]:
            row_data.append(cell.value if cell.value is not None else '')
        result = row_data
        result_type = 'array'
    
    elif request.readMode == 'column':
        if request.columnIndex is None:
            raise HTTPException(status_code=400, detail="列模式需要指定columnIndex")
        col_data = []
        col_idx = request.columnIndex
        if isinstance(col_idx, str):
            col_idx = openpyxl.utils.column_index_from_string(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            col_data.append(row[0].value if row[0].value is not None else '')
        result = col_data
        result_type = 'array'
    
    elif request.readMode == 'range':
        if not request.startCell or not request.endCell:
            raise HTTPException(status_code=400, detail="范围模式需要指定startCell和endCell")
        range_data = []
        for row in ws[f"{request.startCell}:{request.endCell}"]:
            row_data = [cell.value if cell.value is not None else '' for cell in row]
            range_data.append(row_data)
        result = range_data
        result_type = 'matrix'
    
    else:
        raise HTTPException(status_code=400, detail=f"不支持的读取模式: {request.readMode}")
    
    wb.close()
    
    return {'data': result, 'type': result_type}


async def _read_excel_xls(file_path: str, request: ReadExcelRequest):
    """使用xlrd读取.xls文件"""
    wb = xlrd.open_workbook(file_path)
    
    # 选择工作表
    if request.sheetName:
        if request.sheetName not in wb.sheet_names():
            raise HTTPException(status_code=400, detail=f"工作表 '{request.sheetName}' 不存在")
        ws = wb.sheet_by_name(request.sheetName)
    else:
        ws = wb.sheet_by_index(0)
    
    result = None
    result_type = 'unknown'
    
    if request.readMode == 'cell':
        if not request.cellAddress:
            raise HTTPException(status_code=400, detail="单元格模式需要指定cellAddress")
        # 解析单元格地址如 'A1' -> (0, 0)
        col_str, row_str = '', ''
        for c in request.cellAddress:
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        col_idx = _col_letter_to_index(col_str)
        row_idx = int(row_str) - 1
        value = ws.cell_value(row_idx, col_idx)
        result = value if value != '' else ''  # xlrd空单元格返回空字符串
        result_type = 'cell'
    
    elif request.readMode == 'row':
        if request.rowIndex is None:
            raise HTTPException(status_code=400, detail="行模式需要指定rowIndex")
        row_idx = request.rowIndex - 1  # xlrd从0开始
        row_data = ws.row_values(row_idx)
        # xlrd的空单元格已经是空字符串，但为了保险起见还是处理一下
        row_data = [v if v != '' else '' for v in row_data]
        result = row_data
        result_type = 'array'
    
    elif request.readMode == 'column':
        if request.columnIndex is None:
            raise HTTPException(status_code=400, detail="列模式需要指定columnIndex")
        col_idx = request.columnIndex
        if isinstance(col_idx, str):
            col_idx = _col_letter_to_index(col_idx)
        else:
            col_idx = col_idx - 1  # xlrd从0开始
        col_data = ws.col_values(col_idx)
        # xlrd的空单元格已经是空字符串，但为了保险起见还是处理一下
        col_data = [v if v != '' else '' for v in col_data]
        result = col_data
        result_type = 'array'
    
    elif request.readMode == 'range':
        if not request.startCell or not request.endCell:
            raise HTTPException(status_code=400, detail="范围模式需要指定startCell和endCell")
        # 解析范围
        start_col, start_row = _parse_cell_address(request.startCell)
        end_col, end_row = _parse_cell_address(request.endCell)
        range_data = []
        for r in range(start_row, end_row + 1):
            row_data = []
            for c in range(start_col, end_col + 1):
                value = ws.cell_value(r, c)
                row_data.append(value if value != '' else '')
            range_data.append(row_data)
        result = range_data
        result_type = 'matrix'
    
    else:
        raise HTTPException(status_code=400, detail=f"不支持的读取模式: {request.readMode}")
    
    return {'data': result, 'type': result_type}


def _col_letter_to_index(col_str: str) -> int:
    """将列字母转换为索引（A=0, B=1, ...）"""
    result = 0
    for c in col_str.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1


def _parse_cell_address(address: str) -> tuple[int, int]:
    """解析单元格地址，返回(col_idx, row_idx)，从0开始"""
    col_str, row_str = '', ''
    for c in address:
        if c.isalpha():
            col_str += c
        else:
            row_str += c
    return _col_letter_to_index(col_str), int(row_str) - 1


# 提供给执行器使用的函数
def get_asset_path(file_id: str) -> Optional[str]:
    """获取文件路径"""
    if file_id in data_assets:
        return data_assets[file_id]['path']
    return None


def get_asset_by_name(name: str) -> Optional[dict]:
    """通过原始文件名获取资产"""
    for asset in data_assets.values():
        if asset['originalName'] == name:
            return asset
    return None


@router.get("/{file_id}/preview")
async def preview_excel(file_id: str, sheet: Optional[str] = None, max_rows: int = 100, max_cols: int = 50):
    """预览Excel文件数据"""
    if file_id not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[file_id]
    file_path = asset['path']
    is_xls = file_path.lower().endswith('.xls')
    
    try:
        if is_xls:
            return _preview_xls(file_path, sheet, max_rows, max_cols)
        else:
            return _preview_xlsx(file_path, sheet, max_rows, max_cols)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")


def _preview_xlsx(file_path: str, sheet_name: Optional[str], max_rows: int, max_cols: int):
    """预览xlsx文件"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise HTTPException(status_code=400, detail=f"工作表 '{sheet_name}' 不存在")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    data = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols), 1):
        row_data = []
        for cell in row:
            val = cell.value
            row_data.append(str(val) if val is not None else '')
        data.append(row_data)
    
    # 获取实际的行列数
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    
    wb.close()
    
    return {
        'data': data,
        'totalRows': total_rows,
        'totalCols': total_cols,
        'previewRows': len(data),
        'previewCols': max_cols,
    }


def _preview_xls(file_path: str, sheet_name: Optional[str], max_rows: int, max_cols: int):
    """预览xls文件"""
    wb = xlrd.open_workbook(file_path)
    
    if sheet_name:
        if sheet_name not in wb.sheet_names():
            raise HTTPException(status_code=400, detail=f"工作表 '{sheet_name}' 不存在")
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(0)
    
    data = []
    rows_to_read = min(max_rows, ws.nrows)
    cols_to_read = min(max_cols, ws.ncols)
    
    for row_idx in range(rows_to_read):
        row_data = []
        for col_idx in range(cols_to_read):
            val = ws.cell_value(row_idx, col_idx)
            row_data.append(str(val) if val is not None and val != '' else '')
        data.append(row_data)
    
    return {
        'data': data,
        'totalRows': ws.nrows,
        'totalCols': ws.ncols,
        'previewRows': len(data),
        'previewCols': cols_to_read,
    }
